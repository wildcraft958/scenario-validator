"""CH_FB_01 / CH_FB_02 - Functional Block checks (ENCAP functional / TA workbook).

Every scenario ships an ENCAP functional workbook (`ENCAP_Scenario_func_<base>.xlsm`)
used by the dSPACE HIL harness.
  * CH_FB_01 confirms it is present and a valid workbook (OOXML/zip).
  * CH_FB_02 reads the TA scenario sheet's "Set_initial_position" step and checks that the
    object display-switches (columns J-N, `Obj*_DispSW`) and their vertical positions
    (columns O-S, `Obj*_Vertical`) are set for exactly the number of fellows (non-VUT
    actors) in the scenario. It PASSes on a clean match and otherwise asks for a manual
    look - the workbook is macro-driven, so a static read confirms the obvious wiring but
    not the full TA behaviour.
"""
from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

import openpyxl

from ..models import CheckResult, CheckStatus, Config
from .naming import _canonical_base

CATEGORY = "FunctionalBlock"

_DESCRIPTIONS = {
    "CH_FB_01": "ENCAP functional / Test-Automation workbook provided and a valid workbook (ENCAP_Scenario_func)",
    "CH_FB_02": "TA workbook object display-switches (J-N) and positions (O-S) set for the scenario's fellow count",
}


def _make(check_id: str, status: CheckStatus, comment: str = "") -> CheckResult:
    return CheckResult(
        check_id=check_id,
        category=CATEGORY,
        description=_DESCRIPTIONS[check_id],
        status=status,
        comment=comment,
    )


def _locate_workbook(scenario_dir: Path, config: Config) -> tuple[Path | None, str, str]:
    """Return (path_or_None, expected_name, note). Auto-detects a present-but-misnamed file."""
    base = _canonical_base(scenario_dir, config) or scenario_dir.name
    expected = config.functional_file_name(base)
    glob = config.functional_file_pattern.replace("{base}", "*")
    path = scenario_dir / expected
    if path.exists():
        return path, expected, ""
    matches = list(scenario_dir.glob(glob))
    if matches:
        return matches[0], expected, f" (found as '{matches[0].name}', expected '{expected}' - verify base name)"
    return None, expected, ""


def check_fb_01(scenario_dir: Path, config: Config) -> CheckResult:
    """The ENCAP functional workbook must be present and a valid workbook (zip/OOXML)."""
    path, expected, note = _locate_workbook(scenario_dir, config)
    if path is None:
        return _make("CH_FB_01", "FAIL", f"{expected} not found in scenario directory")
    if not zipfile.is_zipfile(path):
        return _make("CH_FB_01", "FAIL", f"{path.name} is not a valid .xlsm workbook (not a zip/OOXML file)")

    result = _make(
        "CH_FB_01",
        "PASS",
        f"{path.name} present and a valid workbook.{note} Column values are checked by CH_FB_02; "
        "confirm it uses the EuroNCAP v4 template.",
    )
    result.source_file = path.name
    return result


def _fellow_names(scenario_dir: Path, config: Config) -> list[str] | None:
    """Non-VUT actor names from the .xosc (the 'fellows'), or None if unreadable."""
    xoscs = list(scenario_dir.glob("*.xosc"))
    if not xoscs:
        return None
    try:
        from ..parsers import xosc as xosc_mod
        root = xosc_mod.load(xoscs[0])
        vut_upper = {n.upper() for n in config.vut_entity_names}
        names = [xosc_mod.get_entity_name(e) for e in xosc_mod.get_entities(root)]
        return [n for n in names if n and n.upper() not in vut_upper]
    except Exception:
        return None


def _as_number(value: Any) -> float | None:
    try:
        return float(str(value))
    except (TypeError, ValueError):
        return None


def _read_init_position_row(path: Path) -> tuple[str, int, list[Any], list[Any]] | None:
    """Find the TA sheet's Set_initial_position step and return
    (sheet, row, DispSW values [J-N], Vertical values [O-S]). None if not found."""
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=False)
    for ws in wb.worksheets:
        header_row = None
        for r in range(1, 8):
            for c in range(1, min(ws.max_column, 40) + 1):
                if ws.cell(row=r, column=c).value == "Obj1_DispSW":
                    header_row = r
                    break
            if header_row:
                break
        if header_row is None:
            continue
        disp_cols = [c for c in range(1, ws.max_column + 1)
                     if str(ws.cell(row=header_row, column=c).value or "").endswith("_DispSW")]
        vert_cols = [c for c in range(1, ws.max_column + 1)
                     if str(ws.cell(row=header_row, column=c).value or "").endswith("_Vertical")]
        set_row = None
        for r in range(header_row, ws.max_row + 1):
            b = ws.cell(row=r, column=2).value
            if isinstance(b, str) and b.startswith("Set_initial_pos"):
                set_row = r
                break
        if set_row is None:
            continue
        disp = [ws.cell(row=set_row, column=c).value for c in disp_cols]
        vert = [ws.cell(row=set_row, column=c).value for c in vert_cols]
        return ws.title, set_row, disp, vert
    return None


def check_fb_02(scenario_dir: Path, config: Config) -> CheckResult:
    """Object display-switches (J-N) and positions (O-S) must be set for the fellow count.

    Conservative by design: PASS only on a clean match (the count of display-enabled objects
    equals the fellow count, those objects have a non-zero position, and every switch is 0 or
    2); anything else is MANUAL_REVIEW with the extracted values, never a brittle hard-FAIL."""
    path, _, note = _locate_workbook(scenario_dir, config)
    if path is None or not zipfile.is_zipfile(path):
        return _make("CH_FB_02", "MANUAL_REVIEW",
                     "ENCAP functional workbook not available - verify the J-N/O-S TA columns manually.")
    fellows = _fellow_names(scenario_dir, config)
    try:
        info = _read_init_position_row(path)
    except Exception as exc:  # noqa: BLE001 - a malformed workbook must not crash the run
        return _make("CH_FB_02", "MANUAL_REVIEW", f"Could not read the TA workbook: {exc}. Verify J-N/O-S manually.")

    if info is None or fellows is None:
        return _make("CH_FB_02", "MANUAL_REVIEW",
                     "Could not locate the Set_initial_position DispSW row or the .xosc actors - verify manually.")

    sheet, row, disp, vert = info
    active = [i for i, d in enumerate(disp) if _as_number(d) == 2.0]
    all_binary = all(_as_number(d) in (0.0, 2.0) for d in disp)
    positions_set = all(
        (_as_number(vert[i]) not in (None, 0.0)) for i in active
    ) if active else False
    n_fellows = len(fellows)

    if len(active) == n_fellows and positions_set and all_binary:
        return _make(
            "CH_FB_02", "PASS",
            f"TA workbook '{sheet}' row {row}: {len(active)} object(s) display-enabled (DispSW=2) "
            f"with positions set, matching {n_fellows} fellow(s) {fellows}.{note}",
        )
    return _make(
        "CH_FB_02", "MANUAL_REVIEW",
        f"TA workbook '{sheet}' row {row}: DispSW(J-N)={disp}, positions(O-S)={vert} for "
        f"{n_fellows} fellow(s) {fellows}. Verify columns J-N are '2' per fellow and O-S hold "
        f"their positions.{note}",
    )


def run_all(scenario_dir: Path, config: Config) -> list[CheckResult]:
    return [
        check_fb_01(scenario_dir, config),
        check_fb_02(scenario_dir, config),
    ]
