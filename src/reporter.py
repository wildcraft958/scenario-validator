"""Excel reporter for EuroNCAP scenario validation results.

Sheet 1: Validation   - one row per check
Sheet 2: Issues Log   - failed checks only
Sheet 3: Run Summary  - aggregate stats and audit metadata
"""
from __future__ import annotations

import logging
from collections.abc import Mapping
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.views import Selection

from .models import CheckResult, SummaryStats
from .rollup import BatchSummaryMeta, ScenarioRow

log = logging.getLogger(__name__)

# Colour palette (ARGB)
_GREEN = "FF92D050"
_RED = "FFFF0000"
_YELLOW = "FFFFC000"
_GREY = "FFD3D3D3"
_BLUE_HDR = "FF4472C4"
_WHITE = "FFFFFFFF"

# Reference-checklist palette - literal ARGB lifted from the reviewer workbook so the
# --checklist export is an exact colour replica. The reference uses the older Office theme
# (accent1 = 5B9BD5), so openpyxl theme references would resolve to a different blue;
# baking the resolved RGB keeps the colours identical.
_REF_LABEL_FILL = PatternFill("solid", fgColor="FF9DC3E6")     # metadata + version + issues headers
_REF_HEADER_FILL = PatternFill("solid", fgColor="FFDEEBF7")    # main checklist header row
_REF_CATEGORY_FILL = PatternFill("solid", fgColor="FFFFFFFF")  # merged category cells (white)
_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Reference column widths for the ChecklistFinal sheet (B-H), used as the default the
# config can still override per column.
_REF_CHECKLIST_WIDTHS = {
    2: 12.140625, 3: 18.7109375, 4: 196.7109375, 5: 12.7109375,
    6: 11.85546875, 7: 11.85546875, 8: 13.0,
}

_STATUS_FILL = {
    "Yes": PatternFill("solid", fgColor=_GREEN),
    "No": PatternFill("solid", fgColor=_RED),
    "NA": PatternFill("solid", fgColor=_GREY),
    "Manual": PatternFill("solid", fgColor=_YELLOW),
}
_STATUS_FONT = {
    "Yes": Font(bold=True, color="FF000000"),
    "No": Font(bold=True, color=_WHITE),
    "NA": Font(bold=False, color="FF000000"),
    "Manual": Font(bold=True, color="FF000000"),
}

# Trust-tier fills for the Automation Level column (green = trust it, yellow = confirm,
# grey = human decides).
_AUTOMATION_FILL = {
    "Fully Automated": PatternFill("solid", fgColor=_GREEN),
    "Partially Automated": PatternFill("solid", fgColor=_YELLOW),
    "Manual": PatternFill("solid", fgColor=_GREY),
}

_VALIDATION_HEADERS = [
    "Check ID",
    "Category",
    "Check name",
    "Result",
    "Comment",
    "Source file",
    "Timestamp",
    "Automation Level",
    "Automation - why",
]
_ISSUE_HEADERS = ["Check ID", "Category", "Issue", "File", "Suggested fix"]
_VALIDATION_COL_CHARS = {1: 16, 2: 16, 3: 52, 4: 10, 5: 60, 6: 22, 7: 20, 8: 20, 9: 50}

# Root batch-summary (Design A: trust dashboard) - one row per scenario.
_SUMMARY_HEADERS = [
    "S/No", "Batch", "Category", "Scenario", "Total", "Automated", "Passed",
    "Failed", "Manual", "NA", "Confidence", "Verdict", "Advice", "Path",
]
_SUMMARY_COL_WIDTHS = {
    1: 6, 2: 16, 3: 18, 4: 46, 5: 8, 6: 11, 7: 9, 8: 8, 9: 9, 10: 6,
    11: 12, 12: 9, 13: 64, 14: 70,
}
# Column numbers are derived from the header order so the writer never carries magic
# indices: rename or reorder a header and the styling follows.
_SUMMARY_COL = {name: i + 1 for i, name in enumerate(_SUMMARY_HEADERS)}
_SUMMARY_LEFT_COLS = {_SUMMARY_COL[n] for n in ("Batch", "Category", "Scenario", "Advice", "Path")}


def _header_row(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    fills = PatternFill("solid", fgColor=_BLUE_HDR)
    font = Font(bold=True, color=_WHITE)
    for col_idx, text in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill = fills
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _result_row(ws, row: int, result: CheckResult) -> None:
    values = result.as_validation_row()
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True)
        if col_idx == 4:
            cell.fill = _STATUS_FILL.get(str(value), PatternFill("solid", fgColor=_GREY))
            cell.font = _STATUS_FONT.get(str(value), Font())
            cell.alignment = Alignment(horizontal="center")
        elif col_idx == 8:
            cell.fill = _AUTOMATION_FILL.get(str(value), PatternFill("solid", fgColor=_GREY))
            cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _set_column_widths(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _auto_row_heights(ws, data_start_row: int, col_char_widths: Mapping[int, float],
                      min_height: float = 15.0, line_height: float = 14.0) -> None:
    """Set row heights so wrapped text is fully visible without double-clicking.

    openpyxl sets wrap_text=True but leaves row height at the Excel default (15pt),
    which clips multi-line content. This function estimates the number of wrapped
    lines in each cell based on text length ÷ column character width and sizes
    the row to fit the tallest cell.
    """
    import math
    for row in ws.iter_rows(min_row=data_start_row):
        max_lines = 1
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            col_idx = cell.column
            col_w = col_char_widths.get(col_idx, 15)
            # Each Excel "character width" unit ≈ 1 character at default font. Count wrapped
            # lines per explicit paragraph (split on newlines) so multi-line comments are not
            # clipped, then sum.
            lines = sum(max(1, math.ceil(len(seg) / max(col_w, 1))) for seg in text.split("\n"))
            max_lines = max(max_lines, max(1, lines))
        height = min_height + (max_lines - 1) * line_height
        ws.row_dimensions[row[0].row].height = height


def write_excel(
    results: list[CheckResult],
    stats: SummaryStats,
    output_path: Path,
) -> Path:
    """Write full Excel report. Returns path of written file."""
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for sheet_name in ("Validation", "Issues Log", "Run Summary"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    # ---- Sheet 1: Validation results ----
    ws = wb.create_sheet("Validation", 0)
    title_cell = ws.cell(row=1, column=1, value=f"EuroNCAP Scenario Validation - {stats.scenario_name}")
    title_cell.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_VALIDATION_HEADERS))

    _header_row(ws, row=3, headers=_VALIDATION_HEADERS)

    data_start = 4
    for i, result in enumerate(results):
        row = data_start + i
        _result_row(ws, row, result)

    _set_column_widths(ws, _VALIDATION_COL_CHARS)
    ws.freeze_panes = "A4"
    # openpyxl leaves the bottom (scrollable) pane's active cell at A1, which sits inside the
    # frozen header region. Excel then scrolls the bottom pane up to row 1 on redraw, painting
    # the title + header rows a SECOND time below the frozen ones (the "duplicate header"
    # artifact). Anchor the bottom-pane selection to the first scrollable cell so it opens clean.
    ws.sheet_view.selection = [Selection(pane="bottomLeft", activeCell="A4", sqref="A4")]
    # Set header rows to fixed heights; auto-size data rows
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[3].height = 28
    _auto_row_heights(ws, data_start_row=4, col_char_widths=_VALIDATION_COL_CHARS)

    # ---- Sheet 2: Issues log ----
    ws_issues = wb.create_sheet("Issues Log")
    _header_row(ws_issues, row=1, headers=_ISSUE_HEADERS)

    row = 2
    for result in results:
        if result.status == "FAIL":
            values = [
                result.check_id,
                result.category,
                result.comment,
                result.source_file,
                result.suggested_fix,
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws_issues.cell(row=row, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True)
            row += 1

    _ISSUES_COL_WIDTHS = {1: 16, 2: 16, 3: 64, 4: 22, 5: 64}
    for col, width in _ISSUES_COL_WIDTHS.items():
        ws_issues.column_dimensions[get_column_letter(col)].width = width
    ws_issues.row_dimensions[1].height = 28
    _auto_row_heights(ws_issues, data_start_row=2, col_char_widths=_ISSUES_COL_WIDTHS)

    # ---- Sheet 3: Run Summary ----
    ws_sum = wb.create_sheet("Run Summary")

    fully = sum(1 for r in results if r.automation_level == "Fully Automated")
    partial = sum(1 for r in results if r.automation_level == "Partially Automated")
    manual_auto = sum(1 for r in results if r.automation_level == "Manual")
    summary_rows = [
        ("Scenario Directory", stats.scenario_dir),
        ("Scenario Name", stats.scenario_name),
        ("Run Timestamp", stats.run_timestamp),
        ("Config Path", stats.config_path),
        ("Protocol Version", stats.protocol_version),
        ("Total Checks", stats.total),
        ("Yes Count", stats.passed),
        ("No Count", stats.failed),
        ("NA Count", stats.na),
        ("Manual Count", stats.manual),
        ("Automatable Pass Rate", f"{stats.pass_rate:.1f}%"),
        ("Automation - Fully", fully),
        ("Automation - Partially", partial),
        ("Automation - Manual", manual_auto),
        ("Final Status", stats.final_status),
        ("CLI Command Used", stats.cli_command),
        ("Failed Check IDs", ", ".join(stats.critical_failures) or "None"),
    ]
    header_fill = PatternFill("solid", fgColor=_BLUE_HDR)
    for row_idx, (label, value) in enumerate(summary_rows, start=2):
        lbl_cell = ws_sum.cell(row=row_idx, column=2, value=label)
        lbl_cell.fill = header_fill
        lbl_cell.font = Font(bold=True, color=_WHITE)

        val_cell = ws_sum.cell(row=row_idx, column=3, value=value)
        if label == "Automatable Pass Rate":
            pct = stats.pass_rate
            val_cell.fill = PatternFill(
                "solid",
                fgColor=_GREEN if pct >= 80 else _YELLOW if pct >= 50 else _RED,
            )
        if label == "Final Status":
            val_cell.fill = PatternFill("solid", fgColor=_GREEN if value == "PASS" else _RED)
            val_cell.font = Font(bold=True, color=_WHITE if value == "FAIL" else "FF000000")
        val_cell.alignment = Alignment(wrap_text=True)

    ws_sum.column_dimensions["B"].width = 22
    ws_sum.column_dimensions["C"].width = 60

    wb.save(output_path)
    log.info("Excel report written to %s", output_path)
    return output_path


def _issue_details(r: CheckResult) -> str:
    """Issues Log 'Details' text for a failed or manual check."""
    body = r.comment or r.description
    if r.result == "Manual":
        return f"{r.check_id} ({r.category}): manual check - verify: {body}"
    text = f"{r.check_id} ({r.category}): {body}"
    if r.suggested_fix and r.suggested_fix != r.comment:
        text = f"{text}  Fix: {r.suggested_fix}"
    return text


def write_reference_checklist(
    results: list[CheckResult],
    stats: SummaryStats,
    output_path: Path,
) -> Path:
    """Write a workbook that replicates the reviewer checklist (Summary / ChecklistFinal /
    Prequisites) - same sheets, column layout and colours - so it drops straight into the
    team's review flow. Self Review carries our verdict as text (like the reviewer file);
    Review1/Review2 stay blank for humans. The Issues Log table is filled from the run: one
    row per failed or manual check. Checkpoints the validator does not compute (MD_06-11)
    export as Manual rows with the reference wording."""
    from .checklist_template import (
        CHECKLIST_COLUMNS,
        CHECKLIST_HEADER_LABELS,
        ISSUES_LOG_COLUMNS,
        MASTER_CHECKLIST,
        PREREQUISITES,
        RELEASE_OPTIONS,
        SELF_REVIEW_OPTIONS,
        SEVERITY_OPTIONS,
        STATUS_OPTIONS,
        SUMMARY_META,
    )

    by_id = {r.check_id: r for r in results}
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ---- Sheet 1: Summary ----
    ws = wb.create_sheet("Summary")
    for i, (label, value) in enumerate(SUMMARY_META, start=4):
        ws.cell(row=i, column=3, value=label).fill = _REF_LABEL_FILL
        ws.cell(row=i, column=4, value=value)
    ws.column_dimensions["C"].width = 16.85546875
    ws.column_dimensions["D"].width = 62.0

    # ---- Sheet 2: ChecklistFinal ----
    ws = wb.create_sheet("ChecklistFinal")
    header_values = {
        "Release": "PriorityA",
        "Protocol Name": stats.protocol_version,
        "Responsible": "",
        "Reviewer": "",
        "Date": stats.run_timestamp,
    }
    for i, label in enumerate(CHECKLIST_HEADER_LABELS, start=2):
        ws.cell(row=i, column=3, value=label).fill = _REF_LABEL_FILL
        ws.cell(row=i, column=4, value=header_values.get(label, ""))

    # -- Main checklist table (the reference six columns) --
    table_header_row = 8
    for col_off, text in enumerate(CHECKLIST_COLUMNS):
        cell = ws.cell(row=table_header_row, column=2 + col_off, value=text)
        cell.fill = _REF_HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER_ALL

    first_data_row = table_header_row + 1
    row = first_data_row
    cat_spans: list[tuple[str, int, int]] = []
    cat_start = first_data_row
    last_cat: str | None = None
    for cat, cid, ref_text in MASTER_CHECKLIST:
        result = by_id.get(cid)
        if cat != last_cat:
            if last_cat is not None:
                cat_spans.append((last_cat, cat_start, row - 1))
            last_cat = cat
            cat_start = row
        cid_cell = ws.cell(row=row, column=3, value=cid)
        cid_cell.alignment = Alignment(horizontal="center", vertical="center")
        cid_cell.border = _BORDER_ALL
        d = ws.cell(row=row, column=4, value=result.description if result else ref_text)
        d.alignment = Alignment(wrap_text=True, vertical="center")
        d.border = _BORDER_ALL
        # Self Review = our verdict, RAG-coloured (Yes=green / No=red / NA=grey / Manual=yellow)
        # so the outcome reads at a glance.
        verdict = result.result if result else "Manual"
        sr = ws.cell(row=row, column=5, value=verdict)
        sr.alignment = Alignment(horizontal="center", vertical="center")
        sr.border = _BORDER_ALL
        sr.fill = _STATUS_FILL.get(verdict, PatternFill("solid", fgColor=_GREY))
        sr.font = _STATUS_FONT.get(verdict, Font())
        # Review1 / Review2 left blank for human reviewers (bordered).
        ws.cell(row=row, column=6).border = _BORDER_ALL
        ws.cell(row=row, column=7).border = _BORDER_ALL
        row += 1
    if last_cat is not None:
        cat_spans.append((last_cat, cat_start, row - 1))
    last_data_row = row - 1

    # Category column: one white merged cell per category, like the reference.
    for cat, start, end in cat_spans:
        for rr in range(start, end + 1):
            cc = ws.cell(row=rr, column=2)
            cc.fill = _REF_CATEGORY_FILL
            cc.border = _BORDER_ALL
        head = ws.cell(row=start, column=2, value=cat)
        head.alignment = Alignment(horizontal="center", vertical="center")
        if end > start:
            ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)

    # Dropdowns: Yes/No on the human review columns, Release on the metadata cell.
    yn = DataValidation(type="list", formula1=f'"{SELF_REVIEW_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(yn)
    yn.add(f"F{first_data_row}:G{last_data_row}")
    rel = DataValidation(type="list", formula1=f'"{RELEASE_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(rel)
    rel.add("D2")

    # Column widths - reference values, still overridable per column from config.
    checklist_widths = {**_REF_CHECKLIST_WIDTHS, **stats.checklist_column_widths}
    for col, width in checklist_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    _auto_row_heights(ws, data_start_row=first_data_row, col_char_widths=checklist_widths)

    # -- Issues Log table (filled from the run: one row per FAIL or MANUAL check) --
    issues = [r for r in results if r.result in ("No", "Manual")]
    issues_header_row = last_data_row + 3
    for col_off, text in enumerate(ISSUES_LOG_COLUMNS):
        cell = ws.cell(row=issues_header_row, column=2 + col_off, value=text)
        cell.fill = _REF_LABEL_FILL
        cell.border = _BORDER_ALL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=(text == "SelfReview Comment")
        )
    ws.row_dimensions[issues_header_row].height = 28.9

    # Keep a few empty template rows even with no issues, so it still reads as a table.
    n_rows = max(len(issues), 5)
    for idx in range(n_rows):
        rr = issues_header_row + 1 + idx
        for col in range(2, 9):
            ws.cell(row=rr, column=col).border = _BORDER_ALL
        if idx < len(issues):
            r = issues[idx]
            # Colour the Sr No cell by severity (No=red / Manual=yellow) so failures stand
            # out from manual-review items at a glance.
            sr_no = ws.cell(row=rr, column=2, value=idx + 1)
            sr_no.alignment = Alignment(horizontal="center", vertical="center")
            sr_no.fill = _STATUS_FILL.get(r.result, PatternFill("solid", fgColor=_GREY))
            sr_no.font = _STATUS_FONT.get(r.result, Font())
            ws.cell(row=rr, column=4, value=_issue_details(r)).alignment = Alignment(
                wrap_text=True, vertical="center"
            )
    last_issue_row = issues_header_row + n_rows

    sev = DataValidation(type="list", formula1=f'"{SEVERITY_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(sev)
    sev.add(f"C{issues_header_row + 1}:C{last_issue_row}")
    stat = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(stat)
    stat.add(f"E{issues_header_row + 1}:E{last_issue_row}")
    _auto_row_heights(ws, data_start_row=issues_header_row + 1, col_char_widths=checklist_widths)

    # ---- Sheet 3: Prequisites ----
    ws = wb.create_sheet("Prequisites")
    for col, (text, horiz) in enumerate([("SL.NO", "center"), ("Rules", "left")], start=1):
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal=horiz, vertical="center")
    for i, rule in enumerate(PREREQUISITES, start=1):
        ws.cell(row=2 + i, column=1, value=i).alignment = Alignment(
            horizontal="center", vertical="top"
        )
        ws.cell(row=2 + i, column=2, value=rule).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
    ws.column_dimensions["A"].width = 5.7109375
    ws.column_dimensions["B"].width = 95.7109375
    _auto_row_heights(ws, data_start_row=3, col_char_widths={1: 6, 2: 96})

    wb.save(output_path)
    log.info("Reference checklist written to %s", output_path)
    return output_path


def _confidence_style(confidence: str) -> tuple[PatternFill, Font]:
    if confidence == "High":
        return PatternFill("solid", fgColor=_GREEN), Font(bold=True)
    if confidence == "Medium":
        return PatternFill("solid", fgColor=_YELLOW), Font(bold=True)
    if confidence == "Low":
        return PatternFill("solid", fgColor=_RED), Font(bold=True, color=_WHITE)
    return PatternFill("solid", fgColor=_GREY), Font()


def _verdict_style(row: ScenarioRow) -> tuple[PatternFill, Font]:
    if row.verdict == "P":
        return PatternFill("solid", fgColor=_GREEN), Font(bold=True)
    if row.verdict == "ERROR":
        return PatternFill("solid", fgColor=_GREY), Font(bold=True)
    # R: red when a real failure is present, amber when it is manual-only.
    if row.failed > 0:
        return PatternFill("solid", fgColor=_RED), Font(bold=True, color=_WHITE)
    return PatternFill("solid", fgColor=_YELLOW), Font(bold=True)


def write_batch_summary(
    rows: list[ScenarioRow],
    meta: BatchSummaryMeta,
    output_path: Path,
) -> Path:
    """Write the root-level trust dashboard: a top box of run-wide figures over a flat
    one-row-per-scenario table (Design A), plus a Skipped & Errors sheet so nothing that
    was not validated can be mistaken for a pass. Lets a reviewer triage a whole batch
    without opening each file."""
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws = wb.create_sheet("Summary")

    # ---- Top box ----
    title = ws.cell(row=1, column=1, value="EuroNCAP Batch Validation Summary")
    title.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_SUMMARY_HEADERS))

    passes = sum(1 for s in rows if s.verdict == "P")
    reviews = sum(1 for s in rows if s.verdict == "R")
    high = sum(1 for s in rows if s.confidence == "High")
    med = sum(1 for s in rows if s.confidence == "Medium")
    low = sum(1 for s in rows if s.confidence == "Low")
    box_rows: list[tuple[str, object]] = [
        ("Root folder", meta.root),
        ("Run timestamp", meta.run_timestamp),
        ("Scenarios discovered", meta.discovered),
        ("Validated", meta.validated),
        ("Skipped (no .xosc)", meta.skipped),
        ("Errored", meta.errored),
        ("Pass (P)", passes),
        ("Review (R)", reviews),
        ("Checks/scenario (max)", meta.checks_per_scenario),
        ("Confidence  High / Med / Low", f"{high} / {med} / {low}"),
    ]
    box_fill = PatternFill("solid", fgColor=_BLUE_HDR)
    br = 2
    for label, value in box_rows:
        lcell = ws.cell(row=br, column=1, value=label)
        lcell.fill = box_fill
        lcell.font = Font(bold=True, color=_WHITE)
        lcell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=br, start_column=1, end_row=br, end_column=3)
        vcell = ws.cell(row=br, column=4, value=value)
        vcell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=br, start_column=4, end_row=br, end_column=len(_SUMMARY_HEADERS))
        br += 1

    # ---- Per-scenario table ----
    header_row = br + 1  # one blank row after the box
    _header_row(ws, row=header_row, headers=_SUMMARY_HEADERS)
    ws.row_dimensions[header_row].height = 26
    data_start = header_row + 1

    for i, srow in enumerate(rows, start=1):
        r = data_start + i - 1
        values = [
            i, srow.batch, srow.category, srow.scenario, srow.total, srow.automated,
            srow.passed, srow.failed, srow.manual, srow.na, srow.confidence,
            srow.verdict, srow.advice, srow.path,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border = _BORDER_ALL
            if col in _SUMMARY_LEFT_COLS:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=(col == _SUMMARY_COL["Advice"])
                )
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        if srow.failed > 0:
            ws.cell(row=r, column=_SUMMARY_COL["Failed"]).font = Font(bold=True, color=_RED)
        conf_cell = ws.cell(row=r, column=_SUMMARY_COL["Confidence"])
        conf_cell.fill, conf_cell.font = _confidence_style(srow.confidence)
        verdict_cell = ws.cell(row=r, column=_SUMMARY_COL["Verdict"])
        verdict_cell.fill, verdict_cell.font = _verdict_style(srow)

    _set_column_widths(ws, _SUMMARY_COL_WIDTHS)
    if rows:
        # Deliberately no frozen panes. Freezing at the header row would also pin the 13-row
        # top box above it (freeze panes always pins a contiguous top block), which eats the
        # screen and makes the sheet feel unscrollable. The sheet opens at A1 with the box
        # visible and scrolls normally.
        _auto_row_heights(ws, data_start_row=data_start, col_char_widths=_SUMMARY_COL_WIDTHS)

    # ---- Sheet 2: Skipped & Errors ----
    ws2 = wb.create_sheet("Skipped & Errors")
    hdr_fill = PatternFill("solid", fgColor=_BLUE_HDR)
    hdr_font = Font(bold=True, color=_WHITE)

    a = ws2.cell(row=1, column=1, value="Skipped - RoadRunner-native export, no .xosc (NOT validated)")
    a.fill, a.font = hdr_fill, hdr_font
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    rr = 2
    for d in meta.incompatible_dirs or ["None"]:
        ws2.cell(row=rr, column=1, value=d)
        rr += 1

    rr += 1
    b = ws2.cell(row=rr, column=1, value="Errors - crashed during validation (re-run these)")
    b.fill, b.font = hdr_fill, hdr_font
    ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
    rr += 1
    eh1 = ws2.cell(row=rr, column=1, value="Scenario")
    eh2 = ws2.cell(row=rr, column=2, value="Error")
    eh1.font = eh2.font = Font(bold=True)
    rr += 1
    for rel, msg in meta.error_details or [("None", "")]:
        ws2.cell(row=rr, column=1, value=rel)
        ws2.cell(row=rr, column=2, value=msg).alignment = Alignment(wrap_text=True)
        rr += 1

    rr += 1
    c = ws2.cell(row=rr, column=1, value="Report write warnings - validation succeeded, the report file did not write")
    c.fill, c.font = hdr_fill, hdr_font
    ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
    rr += 1
    wh1 = ws2.cell(row=rr, column=1, value="Scenario")
    wh2 = ws2.cell(row=rr, column=2, value="Reason")
    wh1.font = wh2.font = Font(bold=True)
    rr += 1
    for rel, msg in meta.report_warnings or [("None", "")]:
        ws2.cell(row=rr, column=1, value=rel)
        ws2.cell(row=rr, column=2, value=msg).alignment = Alignment(wrap_text=True)
        rr += 1

    ws2.column_dimensions["A"].width = 70
    ws2.column_dimensions["B"].width = 70

    wb.save(output_path)
    log.info("Batch summary written to %s", output_path)
    return output_path


def setup_logging(log_path: Path, quiet: bool = False) -> None:
    """Configure file + console logging."""
    fmt = "%(asctime)s %(levelname)-8s %(name)s - %(message)s"
    handlers: list[logging.Handler] = [logging.FileHandler(log_path, encoding="utf-8")]
    if not quiet:
        handlers.append(logging.StreamHandler())
    logging.basicConfig(
        level=logging.INFO,
        format=fmt,
        handlers=handlers,
        force=True,
    )
