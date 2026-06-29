"""Tests for the Eval_Data benchmark features:

  * CH_NM_02 (reference-canonical): the .rrscene name must equal the .rrscenario name.
  * the automation-level registry + the Automation columns on CheckResult.
  * the --checklist reviewer export (Summary / ChecklistFinal / Prequisites).
  * the .rd parser reading the dSPACE ModelDesk RouteSection schema (CH_MD_03 data).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]


@pytest.fixture(scope="module")
def config():
    from src.models import Config
    return Config.load(ROOT / "config.json")


def _write_rd(tmp_path: Path, content: bytes) -> Path:
    p = tmp_path / "x.rd"
    p.write_bytes(content)
    return p


# ---------------------------------------------------------------------------
# CH_NM_02: .rrscene name equals .rrscenario name
# ---------------------------------------------------------------------------

class TestNM02RrsceneMatchesRrscenario:
    def test_matching_names_pass(self, tmp_path):
        base = "AEB_CCFtap_10VUT_30GVT_50Imp"
        (tmp_path / f"{base}.rrscene").write_text("x", encoding="utf-8")
        (tmp_path / f"{base}.rrscenario").write_text("y", encoding="utf-8")
        from src.checks.naming import check_nm_02
        result = check_nm_02(tmp_path)
        assert result.check_id == "CH_NM_02"
        assert result.status == "PASS", result.comment

    def test_mismatched_names_fail(self, tmp_path):
        (tmp_path / "scene_A.rrscene").write_text("x", encoding="utf-8")
        (tmp_path / "scenario_B.rrscenario").write_text("y", encoding="utf-8")
        from src.checks.naming import check_nm_02
        assert check_nm_02(tmp_path).status == "FAIL"

    def test_missing_one_defers_to_manual(self, tmp_path):
        (tmp_path / "only.rrscene").write_text("x", encoding="utf-8")
        from src.checks.naming import check_nm_02
        assert check_nm_02(tmp_path).status == "MANUAL_REVIEW"


# ---------------------------------------------------------------------------
# Automation-level registry
# ---------------------------------------------------------------------------

_ALL_CHECK_IDS = (
    [f"CH_NM_0{i}" for i in range(1, 7)]
    + [f"CH_RD_0{i}" for i in range(1, 7)]
    + [f"CH_SC_{i:02d}" for i in range(1, 23)]
    + [f"CH_MD_{i:02d}" for i in range(1, 12)]
    + ["CH_MR_01", "CH_MR_02", "CH_FB_01", "CH_FB_02"]
)


class TestAutomationRegistry:
    def test_every_check_id_is_classified(self):
        from src.automation import AUTOMATION_LEVEL
        missing = [c for c in _ALL_CHECK_IDS if c not in AUTOMATION_LEVEL]
        assert not missing, f"unclassified check ids: {missing}"

    def test_levels_and_reasons_are_valid(self):
        from src.automation import AUTOMATION_LEVEL
        allowed = {"Fully Automated", "Partially Automated", "Manual"}
        for cid, (level, note) in AUTOMATION_LEVEL.items():
            assert level in allowed, f"{cid} has bad level {level!r}"
            assert note.strip(), f"{cid} has empty reason"

    def test_checkresult_autopopulates_level_and_row(self):
        from src.models import CheckResult
        r = CheckResult(check_id="CH_SC_16", category="Scenario", description="d", status="MANUAL_REVIEW")
        assert r.automation_level == "Partially Automated"
        assert r.automation_note
        # Validation row carries the two new columns at the end.
        assert len(r.as_validation_row()) == 9
        assert r.as_validation_row()[7] == "Partially Automated"


# ---------------------------------------------------------------------------
# Reviewer checklist export
# ---------------------------------------------------------------------------

class TestReferenceChecklistExport:
    def _build(self, tmp_path):
        from src.models import CheckResult, SummaryStats
        from src.reporter import write_reference_checklist
        results = [
            CheckResult(check_id="CH_NM_01", category="Naming", description="naming", status="PASS"),
            CheckResult(check_id="CH_MD_03", category="ModelDesk", description="routes", status="FAIL",
                        comment="only 1 road in the route"),
            CheckResult(check_id="CH_SC_16", category="Scenario", description="impact",
                        status="MANUAL_REVIEW", comment="impact % needs HIL confirmation"),
        ]
        stats = SummaryStats.from_results(results, "scn", "2026-01-01 00:00:00", "proto")
        out = tmp_path / "Review_Checklist.xlsx"
        write_reference_checklist(results, stats, out)
        return openpyxl.load_workbook(out)

    def test_three_reference_sheets(self, tmp_path):
        wb = self._build(tmp_path)
        assert wb.sheetnames == ["Summary", "ChecklistFinal", "Prequisites"]

    def test_columns_are_the_reference_six(self):
        # The replica drops our automation columns - it carries only the reviewer's six.
        from src.checklist_template import CHECKLIST_COLUMNS, ISSUES_LOG_COLUMNS
        assert CHECKLIST_COLUMNS == [
            "Category", "CheckPoint Number", "Check Points", "Self Review", "Review1", "Review2",
        ]
        assert ISSUES_LOG_COLUMNS == [
            "Sr No ", "Severity", "Details", "Status",
            "SelfReview Comment", "R1 Comment", "R2 Comment",
        ]

    def test_table_header_and_verdicts(self, tmp_path):
        wb = self._build(tmp_path)
        ws = wb["ChecklistFinal"]
        headers = [ws.cell(row=8, column=c).value for c in range(2, 8)]
        assert headers == [
            "Category", "CheckPoint Number", "Check Points", "Self Review", "Review1", "Review2",
        ]
        # no automation columns in the replica (column H on the header row is empty)
        assert ws.cell(row=8, column=8).value is None

        verdict = {}
        for r in range(9, ws.max_row + 1):
            cid = ws.cell(row=r, column=3).value
            if cid and str(cid).startswith("CH_"):
                verdict[cid] = ws.cell(row=r, column=5).value
        # full reference list (NM 6 + RD 6 + SC 22 + MD 11 + MR 2 + FB 2)
        assert len(verdict) == 49
        # our computed checks carry their verdict
        assert verdict["CH_NM_01"] == "Yes"
        assert verdict["CH_MD_03"] == "No"
        # a reference-only check we do not compute exports as Manual
        assert verdict["CH_MD_07"] == "Manual"
        assert verdict["CH_FB_02"] == "Manual"

    def test_no_frozen_panes_opens_at_top(self, tmp_path):
        # The reviewer file has no frozen panes; matching it also removes the scroll/
        # duplicate-header artifact (the saved view opens at the top of every sheet).
        wb = self._build(tmp_path)
        for name in wb.sheetnames:
            assert wb[name].freeze_panes in (None, "A1"), name

    def test_issues_log_filled_from_results(self, tmp_path):
        wb = self._build(tmp_path)
        ws = wb["ChecklistFinal"]
        hdr = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "Sr No ")
        assert [ws.cell(hdr, c).value for c in range(2, 9)] == [
            "Sr No ", "Severity", "Details", "Status",
            "SelfReview Comment", "R1 Comment", "R2 Comment",
        ]
        details = {}
        for r in range(hdr + 1, ws.max_row + 1):
            sr_no, text = ws.cell(r, 2).value, ws.cell(r, 4).value
            if sr_no and text:
                details[sr_no] = text
                # Severity (col 3) and Status (col 5) stay blank for the reviewer.
                assert ws.cell(r, 3).value is None
                assert ws.cell(r, 5).value is None
        # one row per FAIL + MANUAL check (CH_MD_03 fail, CH_SC_16 manual)
        assert len(details) == 2
        joined = " ".join(details.values())
        assert "CH_MD_03" in joined and "CH_SC_16" in joined


# ---------------------------------------------------------------------------
# .rd parser - dSPACE ModelDesk schema
# ---------------------------------------------------------------------------

class TestSC22ObstructionScope:
    """CH_SC_22 is N/A when there is no static obstruction; an obstruction's asset path is
    still validated (PASS in the NCAP folder, FAIL outside it)."""

    def _root(self, body: bytes):
        from lxml import etree
        xml = b'<?xml version="1.0"?><OpenSCENARIO><Entities>' + body + b"</Entities></OpenSCENARIO>"
        return etree.parse(__import__("io").BytesIO(xml),
                           etree.XMLParser(no_network=True)).getroot()

    _VUT = b'<ScenarioObject name="VUT"><Vehicle name="VUT" model3d="Vehicles/MyCar.rrvehicle"/></ScenarioObject>'

    def test_no_obstruction_is_na(self, config):
        from src.checks.scenario import check_sc_22
        body = self._VUT + b'<ScenarioObject name="GVT"><Vehicle name="GVT" model3d="NCAP Assets/GVT.rrvehicle"/></ScenarioObject>'
        assert check_sc_22(self._root(body), config).status == "NA"

    def test_obstruction_in_ncap_folder_passes(self, config):
        from src.checks.scenario import check_sc_22
        body = self._VUT + b'<ScenarioObject name="Obstruction1"><Vehicle name="Obstruction1" model3d="NCAP Assets/obs.rrvehicle"/></ScenarioObject>'
        assert check_sc_22(self._root(body), config).status == "PASS"

    def test_obstruction_outside_ncap_folder_fails(self, config):
        from src.checks.scenario import check_sc_22
        body = self._VUT + b'<ScenarioObject name="Obstruction1"><Vehicle name="Obstruction1" model3d="Props/box.rrvehicle"/></ScenarioObject>'
        assert check_sc_22(self._root(body), config).status == "FAIL"


class TestObstructionDetectionGeneralizes:
    """A recognised target (a stationary GVT in CCRs, an SOV) is never an obstruction; a
    generic static object still is - so SC_22/RD_04 applicability generalises beyond the
    turn-across samples that ship in Eval_Data."""

    def _root(self, xml: str):
        import io
        from lxml import etree
        return etree.parse(io.BytesIO(xml.encode()), etree.XMLParser(no_network=True)).getroot()

    def test_stationary_target_not_obstruction_generic_object_is(self, config):
        from src.checks.scenario import _obstruction_entity_names
        xml = (
            '<OpenSCENARIO><Entities>'
            '<ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>'
            '<ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>'
            '<ScenarioObject name="Barrier1"><Vehicle name="Barrier1"/></ScenarioObject>'
            '</Entities><Init><Actions>'
            '<Private entityRef="GVT"><PrivateAction><LongitudinalAction><SpeedAction>'
            '<SpeedActionTarget><AbsoluteTargetSpeed value="0"/></SpeedActionTarget>'
            '</SpeedAction></LongitudinalAction></PrivateAction></Private>'
            '<Private entityRef="Barrier1"><PrivateAction><LongitudinalAction><SpeedAction>'
            '<SpeedActionTarget><AbsoluteTargetSpeed value="0"/></SpeedActionTarget>'
            '</SpeedAction></LongitudinalAction></PrivateAction></Private>'
            '</Actions></Init></OpenSCENARIO>'
        )
        obs = _obstruction_entity_names(self._root(xml), config)
        assert "GVT" not in obs, f"stationary GVT target must not be an obstruction: {obs}"
        assert "Barrier1" in obs, f"a generic static object IS an obstruction: {obs}"


class TestRD04ObstructionPrecondition:
    """CH_RD_04's (0,0,0) origin rule only applies to scenarios with static objects at the
    intersection, applied at the run level where the .xosc entities are known."""

    def test_no_obstruction_scenario_is_na(self):
        from validator import run_validation
        results, _ = run_validation(ROOT / "examples" / "CCFtap", skip_rd=True)
        rd04 = {r.check_id: r for r in results}["CH_RD_04"]
        assert rd04.status == "NA", rd04.comment

    def test_obstruction_scenario_still_evaluated(self):
        from validator import run_validation
        results, _ = run_validation(ROOT / "examples" / "CPNCO", skip_rd=True)
        rd04 = {r.check_id: r for r in results}["CH_RD_04"]
        assert rd04.status in ("PASS", "FAIL"), rd04.comment


class TestFB02TaWorkbook:
    """CH_FB_02 reads the TA workbook Set_initial_position row and matches DispSW/positions
    to the fellow count (PASS on a clean match, MANUAL otherwise - never a brittle FAIL)."""

    def _scenario(self, tmp_path, config, n_active: int):
        base = "AEB_CCFtap_10VUT_30GVT_50Imp"
        (tmp_path / f"{base}.rrscene").write_text("x", encoding="utf-8")
        xosc = (
            '<?xml version="1.0"?><OpenSCENARIO><Entities>'
            '<ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>'
            '<ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>'
            "</Entities></OpenSCENARIO>"
        )
        (tmp_path / f"{base}.xosc").write_text(xosc, encoding="utf-8")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TAsenario"
        headers = {10: "Obj1_DispSW", 11: "Obj2_DispSW", 12: "Obj3_DispSW", 13: "Obj4_DispSW",
                   14: "Obj5_DispSW", 15: "Obj1_Vertical", 16: "Obj2_Vertical", 17: "Obj3_Vertical",
                   18: "Obj4_Vertical", 19: "Obj5_Vertical"}
        for col, name in headers.items():
            ws.cell(row=4, column=col, value=name)
        ws.cell(row=10, column=2, value="Set_initial_position")
        for i in range(5):
            ws.cell(row=10, column=10 + i, value=2 if i < n_active else 0)        # DispSW J-N
            ws.cell(row=10, column=15 + i, value=12.5 if i < n_active else 0)      # Vertical O-S
        wb.save(tmp_path / config.functional_file_name(base))
        return tmp_path

    def test_matching_fellow_count_passes(self, config, tmp_path):
        from src.checks.functional_block import check_fb_02
        d = self._scenario(tmp_path, config, n_active=1)  # 1 fellow (GVT), 1 active object
        result = check_fb_02(d, config)
        assert result.status == "PASS", result.comment

    def test_mismatch_is_manual(self, config, tmp_path):
        from src.checks.functional_block import check_fb_02
        d = self._scenario(tmp_path, config, n_active=2)  # 2 active but only 1 fellow
        assert check_fb_02(d, config).status == "MANUAL_REVIEW"

    def test_missing_workbook_is_manual(self, config, tmp_path):
        from src.checks.functional_block import check_fb_02
        (tmp_path / "AEB_CCFtap_10VUT_30GVT_50Imp.rrscene").write_text("x", encoding="utf-8")
        assert check_fb_02(tmp_path, config).status == "MANUAL_REVIEW"


class TestRdSegmentCounts:
    def test_dspace_routesection_counted(self, tmp_path):
        xml = (
            b'<RoadNetwork xmlns="dSPACE.ModelDesk.RoadGenerator.Data">'
            b'<Route><Name>VUT_Trajectory</Name><Sections>'
            b'<RouteSection/><RouteSection/><RouteSection/>'
            b'</Sections></Route></RoadNetwork>'
        )
        from src.parsers import rd
        d = rd.load(_write_rd(tmp_path, xml))
        assert rd.get_route_segment_counts(d) == [3]
        assert d["routes"][0]["name"] == "VUT_Trajectory"

    def test_generic_road_ids_counted(self, tmp_path):
        xml = b'<Routes><Route name="Ego"><Road id="1"/><Road id="2"/></Route></Routes>'
        from src.parsers import rd
        d = rd.load(_write_rd(tmp_path, xml))
        assert rd.get_route_segment_counts(d) == [2]

    def test_single_segment_route_is_short(self, tmp_path):
        xml = (
            b'<RoadNetwork xmlns="dSPACE.ModelDesk.RoadGenerator.Data">'
            b'<Route><Name>R</Name><Sections><RouteSection/></Sections></Route></RoadNetwork>'
        )
        from src.parsers import rd
        from src.checks.model_desk import check_md_03
        d = rd.load(_write_rd(tmp_path, xml))
        assert check_md_03(d).status == "FAIL"
