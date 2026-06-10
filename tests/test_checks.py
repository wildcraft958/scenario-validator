"""Test suite for EuroNCAP scenario validator.

Tests are organised in three layers:
1. Unit tests for individual parser functions (no external file I/O)
2. Integration tests using real scenario files - place files in tests/scenarios/
3. Negative tests (deliberately broken values must produce FAIL)

Run with:
    python -m pytest tests/ -v
or:
    python -m pytest tests/test_checks.py -v

Integration tests are skipped automatically when scenario files are not present.
To enable them, place real exported scenario files in tests/scenarios/ following
the expected directory structure documented in the README.
"""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

_ROOT = Path(__file__).parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from src.models import Config

# Integration test fixtures - skipped when not present
_SCENARIOS_DIR = Path(__file__).parent / "scenarios"
_XODR_STRAIGHT = _SCENARIOS_DIR / "straight" / "road.xodr"
_XODR_JUNCTION = _SCENARIOS_DIR / "junction" / "road.xodr"
_XOSC_SAMPLE   = _SCENARIOS_DIR / "sample" / "scenario.xosc"

_FIXTURES_AVAILABLE = _XODR_STRAIGHT.exists() and _XOSC_SAMPLE.exists()


def _workbook_bytes() -> bytes:
    """Minimal valid OOXML (zip) so CH_FB_01's zipfile check treats it as a workbook."""
    import io
    import zipfile

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", "<Types/>")
    return buf.getvalue()


@pytest.fixture(scope="session")
def config() -> Config:
    return Config.load()


@pytest.fixture(scope="session")
def xodr_straight(config):
    if not _XODR_STRAIGHT.exists():
        pytest.skip("No straight road fixture - place road.xodr in tests/scenarios/straight/")
    from src.parsers import xodr
    return xodr.load(_XODR_STRAIGHT)


@pytest.fixture(scope="session")
def xodr_junction(config):
    if not _XODR_JUNCTION.exists():
        pytest.skip("No junction fixture - place road.xodr in tests/scenarios/junction/")
    from src.parsers import xodr
    return xodr.load(_XODR_JUNCTION)


@pytest.fixture(scope="session")
def xosc_sample(config):
    if not _XOSC_SAMPLE.exists():
        pytest.skip("No scenario fixture - place scenario.xosc in tests/scenarios/sample/")
    from src.parsers import xosc
    return xosc.load(_XOSC_SAMPLE)


# ============================================================
# Parser unit tests (require fixture files)
# ============================================================

class TestXodrParser:
    def test_get_lane_widths_straight(self, xodr_straight):
        from src.parsers import xodr
        widths = xodr.get_lane_widths(xodr_straight)
        assert len(widths) > 0
        assert all(isinstance(w, float) for w in widths)

    def test_road_count_straight(self, xodr_straight):
        from src.parsers import xodr
        count = xodr.get_road_count(xodr_straight)
        assert count >= 1

    def test_has_junctions_straight(self, xodr_straight):
        from src.parsers import xodr
        assert not xodr.has_junctions(xodr_straight)

    def test_has_junctions_junction(self, xodr_junction):
        from src.parsers import xodr
        assert xodr.has_junctions(xodr_junction)

    def test_road_markings_present(self, xodr_straight):
        from src.parsers import xodr
        markings = xodr.get_road_markings(xodr_straight)
        assert isinstance(markings, list)

    def test_find_disconnected_roads_clean(self, xodr_straight):
        from src.parsers import xodr
        disconnected = xodr.find_disconnected_roads(xodr_straight)
        assert isinstance(disconnected, list)


class TestXodrJunction:
    def test_junction_radii(self, xodr_junction):
        from src.parsers import xodr
        radii = xodr.junction_curvature_radii(xodr_junction)
        assert isinstance(radii, list)

    def test_shoulder_lane_check(self, xodr_junction):
        from src.parsers import xodr
        result = xodr.has_shoulder_lane_at_junction(xodr_junction)
        assert isinstance(result, bool)


class TestXoscParser:
    def test_get_entities(self, xosc_sample):
        from src.parsers import xosc
        entities = xosc.get_entities(xosc_sample)
        assert len(entities) >= 2

    def test_get_parameter_declarations(self, xosc_sample):
        from src.parsers import xosc
        params = xosc.get_parameter_declarations(xosc_sample)
        assert isinstance(params, list)

    def test_get_simulation_time(self, xosc_sample):
        from src.parsers import xosc
        sim_time = xosc.get_simulation_time(xosc_sample)
        assert sim_time is None or isinstance(sim_time, float)

    def test_get_entity_catalog_filepaths(self, xosc_sample):
        from src.parsers import xosc
        paths = xosc.get_entity_catalog_filepaths(xosc_sample)
        assert isinstance(paths, dict)


# ============================================================
# Road check integration tests
# ============================================================

class TestRoadChecks:
    def test_rd_01_lane_width_straight(self, xodr_straight, config):
        from src.checks.road import check_rd_01
        result = check_rd_01(xodr_straight, config)
        assert result.check_id == "CH_RD_01"
        assert result.status in ("PASS", "FAIL")

    def test_rd_02_road_count_straight(self, xodr_straight, config):
        from src.checks.road import check_rd_02
        result = check_rd_02(xodr_straight, config)
        assert result.check_id == "CH_RD_02"

    def test_rd_03_na_for_no_junction(self, xodr_straight, config):
        from src.checks.road import check_rd_03
        result = check_rd_03(xodr_straight, config)
        assert result.status == "NA"

    def test_rd_03_junction_radius(self, xodr_junction, config):
        from src.checks.road import check_rd_03
        result = check_rd_03(xodr_junction, config)
        assert result.check_id == "CH_RD_03"
        assert result.status in ("PASS", "FAIL", "NA")

    def test_rd_04_na_for_no_junction(self, xodr_straight, config):
        from src.checks.road import check_rd_04
        result = check_rd_04(xodr_straight, config)
        assert result.status == "NA"

    def test_rd_06_no_shoulder_straight(self, xodr_straight, config):
        from src.checks.road import check_rd_06
        result = check_rd_06(xodr_straight, config)
        assert result.status == "NA"


# ============================================================
# Scenario check integration tests
# ============================================================

class TestScenarioChecks:
    @pytest.fixture(autouse=True)
    def _skip_if_no_fixture(self):
        if not _XOSC_SAMPLE.exists():
            pytest.skip("No scenario fixture - place scenario.xosc in tests/scenarios/sample/")

    def test_sc_01_parameters(self, xosc_sample, config):
        from src.checks.scenario import check_sc_01
        result = check_sc_01(xosc_sample, config)
        assert result.check_id == "CH_SC_01"
        assert result.status in ("PASS", "FAIL", "MANUAL_REVIEW")

    def test_sc_04_simulation_time(self, xosc_sample, config):
        from src.checks.scenario import check_sc_04
        result = check_sc_04(xosc_sample, config)
        assert result.check_id == "CH_SC_04"
        assert result.status in ("PASS", "FAIL", "NA")

    def test_sc_07_clothoid(self, xosc_sample, config):
        from src.checks.scenario import check_sc_07
        result = check_sc_07(xosc_sample, config)
        assert result.check_id == "CH_SC_07"
        assert result.status in ("NA", "PASS", "FAIL", "MANUAL_REVIEW")

    def test_sc_08_always_manual(self, xosc_sample, config):
        from src.checks.scenario import check_sc_08
        result = check_sc_08(xosc_sample, config)
        assert result.status == "MANUAL_REVIEW"

    def test_sc_22_asset_check(self, xosc_sample, config):
        from src.checks.scenario import check_sc_22
        result = check_sc_22(xosc_sample, config)
        assert result.check_id == "CH_SC_22"
        assert result.status in ("PASS", "FAIL", "MANUAL_REVIEW")


# ============================================================
# Model desk checks
# ============================================================

class TestModelDeskChecks:
    def test_md_01_clean_straight_road(self, xodr_straight, config):
        from src.checks.model_desk import check_md_01
        result = check_md_01(xodr_straight)
        assert result.check_id == "CH_MD_01"
        assert result.status in ("PASS", "FAIL")

    def test_md_05_no_rd_data(self):
        from src.checks.model_desk import check_md_05
        result = check_md_05({"routes": []}, is_junction_scenario=False)
        assert result.check_id == "CH_MD_05"
        assert result.status == "PASS"


# ============================================================
# Negative / deliberate failure tests (no external files needed)
# ============================================================

def _parse_xml(data: bytes):
    """Parse raw XML bytes using lxml's secure parser."""
    import io
    from lxml import etree
    parser = etree.XMLParser(no_network=True, resolve_entities=False, load_dtd=False)
    return etree.parse(io.BytesIO(data), parser).getroot()


class TestNamingChecks:
    def test_nm_01_non_encap_actor_name_fails(self, config, tmp_path):
        """Actor named 'MyCustomCar' (not an EuroNCAP name) should fail CH_NM_01."""
        xosc = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRs_Test" author="Test"/>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
            <ScenarioObject name="MyCustomCar"><Vehicle name="MyCustomCar"/></ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        (tmp_path / "test.xosc").write_bytes(xosc)
        from src.checks.naming import check_nm_01
        result = check_nm_01(tmp_path, config)
        assert result.status == "FAIL"
        assert "MyCustomCar" in result.comment

    def test_nm_01_encap_actors_pass(self, config, tmp_path):
        """Ego + GVT both follow EuroNCAP convention - should pass CH_NM_01."""
        xosc = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRs_Test" author="Test"/>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
            <ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        (tmp_path / "test.xosc").write_bytes(xosc)
        from src.checks.naming import check_nm_01
        result = check_nm_01(tmp_path, config)
        assert result.status == "PASS"

    def test_nm_01_vehicle2_prefix_passes(self, config, tmp_path):
        """Vehicle2 starts with 'Vehicle' prefix (EuroNCAP convention) - should pass."""
        xosc = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCFtap_Test" author="Test"/>
          <Entities>
            <ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>
            <ScenarioObject name="Vehicle2"><Vehicle name="Vehicle2"/></ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        (tmp_path / "test.xosc").write_bytes(xosc)
        from src.checks.naming import check_nm_01
        result = check_nm_01(tmp_path, config)
        assert result.status == "PASS"

    def test_nm_01_no_vut_fails(self, config, tmp_path):
        """No recognised VUT entity should fail CH_NM_01."""
        xosc = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="Test" author="Test"/>
          <Entities>
            <ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        (tmp_path / "test.xosc").write_bytes(xosc)
        from src.checks.naming import check_nm_01
        result = check_nm_01(tmp_path, config)
        assert result.status == "FAIL"

    def test_nm_01_no_xosc_manual_review(self, config, tmp_path):
        """Missing .xosc should return MANUAL_REVIEW, not crash."""
        from src.checks.naming import check_nm_01
        result = check_nm_01(tmp_path, config)
        assert result.status == "MANUAL_REVIEW"


class TestNegativeChecks:
    def test_rd_01_wrong_lane_width(self, config):
        """Lane width 3.0 m should fail CH_RD_01."""
        xml = b"""<?xml version="1.0"?>
        <OpenDRIVE>
          <road id="1" length="200" junction="-1">
            <link/>
            <planView><geometry x="0" y="0" hdg="0" length="200"><line/></geometry></planView>
            <lanes>
              <laneSection s="0">
                <right>
                  <lane id="-1" type="driving">
                    <width sOffset="0" a="3.0" b="0" c="0" d="0"/>
                    <roadMark type="solid"/>
                  </lane>
                </right>
              </laneSection>
            </lanes>
          </road>
        </OpenDRIVE>"""
        root = _parse_xml(xml)
        from src.checks.road import check_rd_01
        result = check_rd_01(root, config)
        assert result.status == "FAIL"

    def test_rd_02_single_road_fails(self, config):
        """Single road segment should fail CH_RD_02."""
        xml = b"""<?xml version="1.0"?>
        <OpenDRIVE>
          <road id="1" length="200" junction="-1">
            <link/>
            <planView><geometry x="0" y="0" hdg="0" length="200"><line/></geometry></planView>
            <lanes><laneSection s="0"><right><lane id="-1" type="driving">
              <width sOffset="0" a="3.5" b="0" c="0" d="0"/>
            </lane></right></laneSection></lanes>
          </road>
        </OpenDRIVE>"""
        root = _parse_xml(xml)
        from src.checks.road import check_rd_02
        result = check_rd_02(root, config)
        assert result.status == "FAIL"

    def test_sc_04_speed_dependent_below_minimum(self, config):
        """Simulation time 25s with VUT at 30 km/h (8.33 m/s) should fail - below 35s minimum."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRs_Test" author="Test"/>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="Ego">
                <PrivateAction>
                  <LongitudinalAction>
                    <SpeedAction><SpeedActionTarget>
                      <AbsoluteTargetSpeed value="8.33"/>
                    </SpeedActionTarget></SpeedAction>
                  </LongitudinalAction>
                </PrivateAction>
              </Private>
            </Actions>
          </Init>
          <Storyboard>
            <StopTrigger>
              <ConditionGroup>
                <Condition name="Stop" delay="0" conditionEdge="none">
                  <ByValueCondition>
                    <SimulationTimeCondition value="25" rule="greaterThan"/>
                  </ByValueCondition>
                </Condition>
              </ConditionGroup>
            </StopTrigger>
          </Storyboard>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_04
        result = check_sc_04(root, config)
        assert result.status == "FAIL", f"Expected FAIL for 25s at 30 km/h, got {result.status}: {result.comment}"

    def test_sc_04_speed_dependent_in_range(self, config):
        """Simulation time 50s with VUT at 30 km/h should pass (35-60s band)."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRs_Test" author="Test"/>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="Ego">
                <PrivateAction>
                  <LongitudinalAction>
                    <SpeedAction><SpeedActionTarget>
                      <AbsoluteTargetSpeed value="8.33"/>
                    </SpeedActionTarget></SpeedAction>
                  </LongitudinalAction>
                </PrivateAction>
              </Private>
            </Actions>
          </Init>
          <Storyboard>
            <StopTrigger>
              <ConditionGroup>
                <Condition name="Stop" delay="0" conditionEdge="none">
                  <ByValueCondition>
                    <SimulationTimeCondition value="50" rule="greaterThan"/>
                  </ByValueCondition>
                </Condition>
              </ConditionGroup>
            </StopTrigger>
          </Storyboard>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_04
        result = check_sc_04(root, config)
        assert result.status == "PASS", f"Expected PASS for 50s at 30 km/h, got {result.status}: {result.comment}"

    def test_sc_04_out_of_range_time(self, config):
        """Simulation time of 200s should fail CH_SC_04."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="Test" author="Test"/>
          <Storyboard>
            <StopTrigger>
              <ConditionGroup>
                <Condition name="StopCond" delay="0" conditionEdge="none">
                  <ByValueCondition>
                    <SimulationTimeCondition value="200" rule="greaterThan"/>
                  </ByValueCondition>
                </Condition>
              </ConditionGroup>
            </StopTrigger>
          </Storyboard>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_04
        result = check_sc_04(root, config)
        assert result.status == "FAIL"

    def test_sc_14_vut_never_included_as_static(self, config):
        """VUT must never be checked as a static target even if its name contains 'obstruction'."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="Test"/>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
            <ScenarioObject name="GVTs"><Vehicle name="GVTs"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="Ego">
                <PrivateAction><LongitudinalAction><SpeedAction><SpeedActionTarget>
                  <AbsoluteTargetSpeed value="20"/>
                </SpeedActionTarget></SpeedAction></LongitudinalAction></PrivateAction>
              </Private>
              <Private entityRef="GVTs">
                <PrivateAction><LongitudinalAction><SpeedAction><SpeedActionTarget>
                  <AbsoluteTargetSpeed value="0"/>
                </SpeedActionTarget></SpeedAction></LongitudinalAction></PrivateAction>
              </Private>
            </Actions>
          </Init>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_14
        result = check_sc_14(root, config)
        # GVTs matches "GVTs" pattern → PASS (0 m/s). VUT "Ego" must not be evaluated.
        assert result.status in ("PASS", "NA")

    def test_sc_14_non_zero_speed_fails(self, config):
        """Static target with 5 m/s should fail CH_SC_14."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRs_Test"/>
          <Entities>
            <ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>
            <ScenarioObject name="GVTs"><Vehicle name="GVTs"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="VUT">
                <PrivateAction>
                  <LongitudinalAction>
                    <SpeedAction><SpeedActionTarget>
                      <AbsoluteTargetSpeed value="20"/>
                    </SpeedActionTarget></SpeedAction>
                  </LongitudinalAction>
                </PrivateAction>
              </Private>
              <Private entityRef="GVTs">
                <PrivateAction>
                  <LongitudinalAction>
                    <SpeedAction><SpeedActionTarget>
                      <AbsoluteTargetSpeed value="5"/>
                    </SpeedActionTarget></SpeedAction>
                  </LongitudinalAction>
                </PrivateAction>
              </Private>
            </Actions>
          </Init>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_14
        result = check_sc_14(root, config)
        assert result.status == "FAIL"

    def test_sc_02_accepts_lane_position(self, config):
        """VUT placed via LanePosition (not WorldPosition) must PASS CH_SC_02, not false-FAIL.

        Regression for real Vector NCAP scenarios that position actors with
        <LanePosition> inside the Init TeleportAction.
        """
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRs_Test"/>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
          </Entities>
          <Storyboard><Init><Actions>
            <Private entityRef="Ego">
              <PrivateAction><TeleportAction><Position>
                <LanePosition roadId="0" laneId="-1" s="50" offset="0"/>
              </Position></TeleportAction></PrivateAction>
            </Private>
          </Actions></Init></Storyboard>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_02
        result = check_sc_02(root, config)
        assert result.status == "PASS", f"LanePosition should pass, got {result.status}: {result.comment}"

    def test_identify_vut_does_not_substring_match(self, config):
        """Regression: entity named 'GVT_Vehicle_1' must NOT be identified as VUT.
        Before fix, _identify_vut used 'in' (substring), so 'Vehicle' in 'GVT_Vehicle_1'
        returned True, mis-identifying the target. Now uses exact match."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="Test"/>
          <Entities>
            <ScenarioObject name="GVT_Vehicle_1"><Vehicle name="GVT_Vehicle_1"/></ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_02
        result = check_sc_02(root, config)
        assert result.status == "MANUAL_REVIEW", (
            f"Expected MANUAL_REVIEW (no VUT found), got {result.status}: {result.comment}"
        )

    def test_sc_16_na_without_scenario_tag(self, config):
        """CH_SC_16 returns NA when no scenario tag is detectable (RR version string in FileHeader)."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="Exported from RoadRunner version R2025b" author="RR"/>
          <Entities>
            <ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>
            <ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_16
        result = check_sc_16(root, config)
        assert result.status == "NA", f"Expected NA without tag, got {result.status}: {result.comment}"

    def test_sc_16_activated_by_scenario_tag(self, config):
        """CH_SC_16 runs (not NA) when scenario_tag is passed even though FileHeader has no prefix."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="Exported from RoadRunner version R2025b" author="RR"/>
          <Entities>
            <ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>
            <ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="VUT">
                <PrivateAction><TeleportAction><Position>
                  <WorldPosition x="0" y="0" z="0" h="0"/>
                </Position></TeleportAction></PrivateAction>
              </Private>
              <Private entityRef="GVT">
                <PrivateAction><TeleportAction><Position>
                  <WorldPosition x="20" y="0" z="0" h="0"/>
                </Position></TeleportAction></PrivateAction>
              </Private>
            </Actions>
          </Init>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_16
        result = check_sc_16(root, config, scenario_tag="CCFhol")
        assert result.status != "NA", f"Expected PASS/FAIL/MANUAL_REVIEW, got NA: {result.comment}"

    def test_sc_17_activated_by_scenario_tag(self, config):
        """CH_SC_17 runs (not NA) when scenario_tag='CCRs' is passed explicitly."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="Exported from RoadRunner version R2025b" author="RR"/>
          <Entities>
            <ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>
            <ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="VUT">
                <PrivateAction><TeleportAction><Position>
                  <WorldPosition x="0" y="0" z="0" h="0"/>
                </Position></TeleportAction></PrivateAction>
              </Private>
              <Private entityRef="GVT">
                <PrivateAction><TeleportAction><Position>
                  <WorldPosition x="20" y="0" z="0" h="0"/>
                </Position></TeleportAction></PrivateAction>
              </Private>
            </Actions>
          </Init>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_17
        result = check_sc_17(root, config, scenario_tag="CCRs")
        assert result.status != "NA", f"Expected PASS/FAIL/MANUAL_REVIEW, got NA: {result.comment}"

    def test_sc_18_checks_speed_when_tag_provided(self, config):
        """CH_SC_18 returns PASS when VUT speed is in protocol range and scenario_tag is supplied."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="Exported from RoadRunner version R2025b" author="RR"/>
          <Entities>
            <ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="VUT">
                <PrivateAction><LongitudinalAction><SpeedAction><SpeedActionTarget>
                  <AbsoluteTargetSpeed value="13.89"/>
                </SpeedActionTarget></SpeedAction></LongitudinalAction></PrivateAction>
              </Private>
            </Actions>
          </Init>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_18
        # CCRs vut_speed_range_kmh = [10, 80]; 13.89 m/s = 50 km/h → in range
        result = check_sc_18(root, config, scenario_tag="CCRs")
        assert result.status == "PASS", f"Expected PASS for 50 km/h CCRs, got {result.status}: {result.comment}"

    def test_sc_18_fails_out_of_range_with_tag(self, config):
        """CH_SC_18 returns FAIL when VUT speed exceeds protocol range and scenario_tag is supplied."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="Exported from RoadRunner version R2025b" author="RR"/>
          <Entities>
            <ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="VUT">
                <PrivateAction><LongitudinalAction><SpeedAction><SpeedActionTarget>
                  <AbsoluteTargetSpeed value="41.67"/>
                </SpeedActionTarget></SpeedAction></LongitudinalAction></PrivateAction>
              </Private>
            </Actions>
          </Init>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_18
        # CCRs vut_speed_range_kmh = [10, 80]; 41.67 m/s = 150 km/h → out of range
        result = check_sc_18(root, config, scenario_tag="CCRs")
        assert result.status == "FAIL", f"Expected FAIL for 150 km/h CCRs, got {result.status}: {result.comment}"

    def test_sc_06_north_heading_vut_passes(self, config):
        """VUT heading 90° (north) should PASS — cardinal axis, any world direction is valid."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CPNCO_Test" author="Test"/>
          <Entities>
            <ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="VUT">
                <PrivateAction><TeleportAction><Position>
                  <WorldPosition x="620" y="-8" z="0" h="1.5708"/>
                </Position></TeleportAction></PrivateAction>
              </Private>
            </Actions>
          </Init>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_06
        result = check_sc_06(root, config)
        assert result.status == "PASS", (
            f"VUT heading 90° (north, axis-aligned) should PASS. Got {result.status}: {result.comment}"
        )

    def test_sc_06_diagonal_heading_fails(self, config):
        """VUT heading 45° (diagonal) should FAIL — not aligned to any cardinal axis."""
        import math
        diagonal_rad = math.radians(45)
        xml = f"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="Diag_Test" author="Test"/>
          <Entities>
            <ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="VUT">
                <PrivateAction><TeleportAction><Position>
                  <WorldPosition x="0" y="0" z="0" h="{diagonal_rad}"/>
                </Position></TeleportAction></PrivateAction>
              </Private>
            </Actions>
          </Init>
        </OpenSCENARIO>""".encode()
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_06
        result = check_sc_06(root, config)
        assert result.status == "FAIL", (
            f"VUT heading 45° (off-axis) should FAIL. Got {result.status}: {result.comment}"
        )

    def test_rd_04_road_at_origin_passes(self, config):
        """Leftmost road starting at (0, 0) should PASS CH_RD_04 (position-only check)."""
        xml = b"""<?xml version="1.0"?>
        <OpenDRIVE>
          <junction id="1" name="J1"/>
          <road id="1" length="200" junction="-1">
            <link><successor elementId="1" elementType="junction"/></link>
            <planView><geometry x="0" y="0" hdg="1.5708" length="200"><line/></geometry></planView>
            <lanes><laneSection s="0"><right><lane id="-1" type="driving">
              <width sOffset="0" a="3.5" b="0" c="0" d="0"/>
            </lane></right></laneSection></lanes>
          </road>
        </OpenDRIVE>"""
        root = _parse_xml(xml)
        from src.checks.road import check_rd_04
        # junction_scenario_prefixes must include this scenario tag; use CP prefix
        result = check_rd_04(root, config, scenario_tag="CPNCO")
        assert result.status == "PASS", (
            f"Road at origin (0,0) should PASS regardless of heading. Got {result.status}: {result.comment}"
        )

    def test_rd_04_road_off_origin_fails(self, config):
        """Leftmost road starting at (598, 0) should FAIL CH_RD_04 (not at origin)."""
        xml = b"""<?xml version="1.0"?>
        <OpenDRIVE>
          <junction id="1" name="J1"/>
          <road id="1" length="200" junction="-1">
            <link><successor elementId="1" elementType="junction"/></link>
            <planView><geometry x="598" y="0" hdg="3.14159" length="200"><line/></geometry></planView>
            <lanes><laneSection s="0"><right><lane id="-1" type="driving">
              <width sOffset="0" a="3.5" b="0" c="0" d="0"/>
            </lane></right></laneSection></lanes>
          </road>
        </OpenDRIVE>"""
        root = _parse_xml(xml)
        from src.checks.road import check_rd_04
        result = check_rd_04(root, config, scenario_tag="CPNCO")
        assert result.status == "FAIL", (
            f"Road starting at x=598 should FAIL (not at origin). Got {result.status}: {result.comment}"
        )

    def test_sc_22_catalog_reference_ncap_passes(self, config):
        """CatalogReference with NCAP catalog name should PASS CH_SC_22."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <Entities>
            <ScenarioObject name="VUT">
              <Vehicle name="VUT" vehicleCategory="car">
                <Properties><Property name="model3d" value="Vehicles/MyCar.rrvehicle"/></Properties>
              </Vehicle>
            </ScenarioObject>
            <ScenarioObject name="EPTc">
              <Pedestrian name="EPTc" model="adult">
                <CatalogReference catalogName="NCAP Assets/PedestrianCatalog" entryName="adult_pedestrian"/>
              </Pedestrian>
            </ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_22
        result = check_sc_22(root, config)
        assert result.status == "PASS", f"NCAP catalogName should PASS. Got: {result.comment}"
        assert "[catalog]" in result.comment

    def test_sc_22_catalog_reference_non_ncap_fails(self, config):
        """CatalogReference with non-NCAP catalog name should FAIL CH_SC_22."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <Entities>
            <ScenarioObject name="VUT">
              <Vehicle name="VUT" vehicleCategory="car">
                <Properties><Property name="model3d" value="Vehicles/MyCar.rrvehicle"/></Properties>
              </Vehicle>
            </ScenarioObject>
            <ScenarioObject name="GVT">
              <Vehicle name="GVT" vehicleCategory="car">
                <CatalogReference catalogName="CustomLib" entryName="sedan"/>
              </Vehicle>
            </ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_22
        result = check_sc_22(root, config)
        assert result.status == "FAIL", f"Non-NCAP catalog should FAIL. Got: {result.comment}"
        assert "CustomLib" in result.comment

    def test_sc_07_rht_handedness_inverts_direction(self, config):
        """Positive net heading change maps to Nearside in RHT (opposite to LHT)."""
        from src.parsers import xosc as xosc_parser
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <Entities>
            <ScenarioObject name="VUT"><Vehicle name="VUT"/></ScenarioObject>
          </Entities>
          <Storyboard>
            <Init>
              <Actions>
                <Private entityRef="VUT">
                  <PrivateAction>
                    <RoutingAction>
                      <FollowTrajectoryAction>
                        <Trajectory name="t" closed="false">
                          <Shape>
                            <Polyline>
                              <Vertex time="0.0"><Position><WorldPosition x="0" y="0" h="0.0"/></Position></Vertex>
                              <Vertex time="1.0"><Position><WorldPosition x="8" y="0.5" h="0.1"/></Position></Vertex>
                              <Vertex time="2.0"><Position><WorldPosition x="15" y="2.0" h="0.2"/></Position></Vertex>
                              <Vertex time="3.0"><Position><WorldPosition x="20" y="4.5" h="0.35"/></Position></Vertex>
                              <Vertex time="4.0"><Position><WorldPosition x="22" y="7.0" h="0.5"/></Position></Vertex>
                            </Polyline>
                          </Shape>
                        </Trajectory>
                      </FollowTrajectoryAction>
                    </RoutingAction>
                  </PrivateAction>
                </Private>
              </Actions>
            </Init>
          </Storyboard>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        _, direction_lht = xosc_parser.get_polyline_part2_radius(root, "VUT", handedness="LHT")
        assert direction_lht == "Farside", f"LHT: positive dh should be Farside, got {direction_lht}"
        _, direction_rht = xosc_parser.get_polyline_part2_radius(root, "VUT", handedness="RHT")
        assert direction_rht == "Nearside", f"RHT: positive dh should be Nearside, got {direction_rht}"

    def test_sc_22_sov_exempt_from_ncap_folder(self, config):
        """SOV may be 'a GVT or a real vehicle' per protocol — exempt from NCAP-folder rule."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <Entities>
            <ScenarioObject name="VUT">
              <Vehicle name="VUT" vehicleCategory="car" model3d="Vehicles/MyCar.rrvehicle"/>
            </ScenarioObject>
            <ScenarioObject name="GVT">
              <Vehicle name="GVT" vehicleCategory="car" model3d="NCAP Assets/NCAP_GVT.rrvehicle"/>
            </ScenarioObject>
            <ScenarioObject name="SOV">
              <Vehicle name="SOV" vehicleCategory="car" model3d="Vehicles/RealCar.rrvehicle"/>
            </ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.scenario import check_sc_22
        result = check_sc_22(root, config)
        assert result.status == "PASS", f"SOV with real-vehicle path must be exempt. Got: {result.comment}"
        assert "exempt" in result.comment

    def test_catalog_filepaths_first_catalog_ref_wins(self, config):
        """With multiple CatalogReference elements, the FIRST one must be returned."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <Entities>
            <ScenarioObject name="GVT">
              <Vehicle name="GVT" vehicleCategory="car">
                <CatalogReference catalogName="NCAP_Catalog" entryName="first"/>
                <CatalogReference catalogName="OtherLib" entryName="second"/>
              </Vehicle>
            </ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.parsers import xosc as xosc_parser
        result = xosc_parser.get_entity_catalog_filepaths(root)
        assert result["GVT"] == ("NCAP_Catalog/first", "catalog"), result

    def test_paths_intersect_geometry(self):
        """Crossing polylines intersect; parallel ones do not."""
        from src.geometry import paths_intersect
        a = [{"x": 0, "y": 0}, {"x": 10, "y": 0}]
        b = [{"x": 5, "y": -5}, {"x": 5, "y": 5}]
        hit = paths_intersect(a, b)
        assert hit is not None and abs(hit[0] - 5) < 1e-9 and abs(hit[1]) < 1e-9
        c = [{"x": 0, "y": 1}, {"x": 10, "y": 1}]
        assert paths_intersect(a, c) is None
        assert paths_intersect([], b) is None

    def test_impact_estimate_headon_half_offset(self):
        """Head-on with 0.9 m lateral offset between 1.8 m-wide cars → ~50% width overlap."""
        from src.geometry import estimate_trajectory_impact
        bbox = (0.0, 0.0, 4.5, 1.8)
        # VUT eastbound at y=0, target westbound at y=0.9, closing at 20 m/s
        vut = [{"time": t, "x": 0 + 10.0 * t, "y": 0.0, "h": 0.0} for t in range(0, 11)]
        import math
        tgt = [{"time": t, "x": 100 - 10.0 * t, "y": 0.9, "h": math.pi} for t in range(0, 11)]
        est = estimate_trajectory_impact(vut, tgt, bbox, bbox, target_category="Vehicle")
        assert est is not None and est.contact
        assert abs(est.width_overlap_pct - 50.0) < 2.0, est
        assert abs(est.rel_heading_deg - 180.0) < 1.0

    def test_impact_estimate_headon_dead_centre(self):
        """Dead-centre head-on → 100% overlap (the CCFhol/CCFhos design-flaw signature)."""
        from src.geometry import estimate_trajectory_impact
        import math
        bbox = (0.0, 0.0, 4.5, 1.8)
        vut = [{"time": t, "x": 10.0 * t, "y": 0.0, "h": 0.0} for t in range(0, 11)]
        tgt = [{"time": t, "x": 100 - 10.0 * t, "y": 0.0, "h": math.pi} for t in range(0, 11)]
        est = estimate_trajectory_impact(vut, tgt, bbox, bbox, target_category="Vehicle")
        assert est is not None and est.contact
        assert est.width_overlap_pct > 95.0, est

    def test_impact_estimate_no_contact_reports_min_gap(self):
        """Parallel same-direction paths never meet → contact=False with min gap."""
        from src.geometry import estimate_trajectory_impact
        bbox = (0.0, 0.0, 4.5, 1.8)
        vut = [{"time": t, "x": 10.0 * t, "y": 0.0, "h": 0.0} for t in range(0, 11)]
        tgt = [{"time": t, "x": 10.0 * t, "y": 5.0, "h": 0.0} for t in range(0, 11)]
        est = estimate_trajectory_impact(vut, tgt, bbox, bbox, target_category="Vehicle")
        assert est is not None and not est.contact
        assert est.min_gap_m is not None and 2.0 < est.min_gap_m < 4.0, est

    def test_impact_estimate_pedestrian_front_position(self):
        """Pedestrian crossing timed to be 0.7 m off-centre when the VUT front arrives
        → ~11% from one edge / ~89% from the other."""
        from src.geometry import estimate_trajectory_impact
        import math
        vut_bbox = (0.0, 0.0, 4.5, 1.8)
        ped_bbox = (0.0, 0.0, 0.5, 0.5)
        vut = [{"time": t, "x": 10.0 * t, "y": 0.0, "h": 0.0} for t in range(0, 11)]
        # Pedestrian walks +y, crossing x=52.95: VUT front (x+2.25) arrives there at t=5.07.
        # Tune start so pedestrian centre is at y=-0.7 when crossing the front plane.
        ped = [{"time": t, "x": 52.95, "y": -0.7 - 1.0 * (5.07 - t), "h": math.pi / 2} for t in range(0, 11)]
        est = estimate_trajectory_impact(vut, ped, vut_bbox, ped_bbox, target_category="Pedestrian")
        assert est is not None and est.contact, est
        near_edge = min(est.front_pos_left_pct, est.front_pos_right_pct)
        assert abs(near_edge - 11.0) < 5.0, est

    def test_real_cpta_impact_estimate_passes(self, config):
        """Real CPTA export: estimate ~8% vs expected 10% ±5 → PASS."""
        import pathlib
        xosc_path = pathlib.Path("examples/CPTA/AEB_CPTAno_10VUT_5EPTa_10Imp.xosc")
        if not xosc_path.exists():
            pytest.skip("CPTA example not present")
        from src.parsers import xosc as xosc_parser
        from src.checks.scenario import check_sc_16
        root = xosc_parser.load(xosc_path)
        result = check_sc_16(root, config, scenario_tag="CPTA")
        assert result.status == "PASS", result.comment
        assert "Geometric impact estimate" in result.comment

    def test_config_friendly_error_on_broken_json(self, tmp_path):
        """A trailing comma must produce a ConfigError naming the line, not a stack trace."""
        from src.models import Config, ConfigError
        bad = tmp_path / "broken.json"
        bad.write_text('{\n  "lane_width_m": 3.5,\n}')
        with pytest.raises(ConfigError) as exc:
            Config.load(bad)
        assert "line" in str(exc.value)
        assert "trailing" in str(exc.value).lower() or "comma" in str(exc.value).lower()

    def test_config_friendly_error_on_bad_type(self, tmp_path):
        """A wrong value type must produce a ConfigError naming the key."""
        import json as jsonlib
        from src.models import Config, ConfigError
        raw = jsonlib.loads((Path("config.json")).read_text())
        raw["lane_width_m"] = "not-a-number"
        bad = tmp_path / "badtype.json"
        bad.write_text(jsonlib.dumps(raw))
        with pytest.raises(ConfigError) as exc:
            Config.load(bad)
        assert "lane_width_m" in str(exc.value)

    def test_excel_config_round_trips(self, tmp_path):
        """config.xlsx generated from config.json must load to the identical Config."""
        import sys as _sys
        _sys.path.insert(0, str(Path("tools").resolve()))
        import json as jsonlib
        from make_config_xlsx import build_workbook
        from src.models import Config
        raw = jsonlib.loads(Path("config.json").read_text())
        for key in [k for k in raw if k.startswith("_")]:
            raw.pop(key)
        out = tmp_path / "config.xlsx"
        build_workbook(raw).save(out)
        a = Config.load(Path("config.json")).model_dump()
        b = Config.load(out).model_dump()
        a["naming_convention"] = sorted(a["naming_convention"]["valid_prefixes"])
        b["naming_convention"] = sorted(b["naming_convention"]["valid_prefixes"])
        assert a == b

    def test_real_ccfhol_impact_estimate_flags_dead_centre(self, config):
        """Real CCFhol export collides dead-centre (100%) but protocol says 50% → FAIL."""
        import pathlib
        xosc_path = pathlib.Path("examples/CCFhol/AEB_CCFhol_30VUT_50GVT_50Imp.xosc")
        if not xosc_path.exists():
            pytest.skip("CCFhol example not present")
        from src.parsers import xosc as xosc_parser
        from src.checks.scenario import check_sc_16
        root = xosc_parser.load(xosc_path)
        result = check_sc_16(root, config, scenario_tag="CCFhol")
        assert result.status == "FAIL", result.comment
        assert "100.0%" in result.comment

    def test_nm_03_optional_file_absent_still_passes(self, config, tmp_path):
        """NM_03 must PASS even when optional catalog files are absent."""
        # Create all required files (7 extensions + TA.xml)
        base = "AEB_CCFhol_30VUT_50GVT_50Imp"
        for ext in config.required_file_extensions:
            (tmp_path / f"{base}{ext}").write_text("dummy")
        for standalone in config.required_standalone_files:
            (tmp_path / standalone).write_text("dummy")
        for _role, expected, _glob, required in config.associated_files(base):
            if required:
                (tmp_path / expected).write_bytes(_workbook_bytes())
        # Do NOT create any of the optional_standalone_files (VehicleCatalog.xosc etc.)
        from src.checks.naming import check_nm_03
        result = check_nm_03(tmp_path, config)
        assert result.status == "PASS", f"All required files present; optional absent should PASS. Got: {result.comment}"
        if config.optional_standalone_files:
            assert "not required" in result.comment or "Optional" in result.comment, (
                "Comment should mention optional files status"
            )


# ============================================================
# Geometry tests (no external files needed)
# ============================================================

class TestModelReviewChecks:
    def _braking_xosc(self, decel_value: str) -> bytes:
        return f"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRb_Test" author="Test"/>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
            <ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>
          </Entities>
          <Storyboard>
            <Story name="GVT_Braking">
              <Act name="Act">
                <ManeuverGroup name="MG" maximumExecutionCount="1">
                  <Actors selectTriggeringEntities="false">
                    <EntityRef entityRef="GVT"/>
                  </Actors>
                  <Maneuver name="M">
                    <Event name="E" priority="overwrite">
                      <Action name="A">
                        <PrivateAction>
                          <LongitudinalAction>
                            <SpeedAction>
                              <SpeedActionDynamics dynamicsDimension="rate" dynamicsShape="linear" value="{decel_value}"/>
                              <SpeedActionTarget>
                                <AbsoluteTargetSpeed value="0"/>
                              </SpeedActionTarget>
                            </SpeedAction>
                          </LongitudinalAction>
                        </PrivateAction>
                      </Action>
                      <StartTrigger/>
                    </Event>
                  </Maneuver>
                </ManeuverGroup>
                <StartTrigger/>
              </Act>
            </Story>
            <StopTrigger/>
          </Storyboard>
        </OpenSCENARIO>""".encode()

    def test_mr_02_na_for_non_braking_scenario(self, config):
        """Scenario with no linear-rate decel action -> NA."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRs_Test"/>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
            <ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>
          </Entities>
          <Storyboard><StopTrigger/></Storyboard>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.model_review import check_mr_02
        result = check_mr_02(root, config)
        assert result.status == "NA"

    def test_mr_02_pass_correct_decel(self, config):
        """GVT decel = 4.0 m/s² (expected) -> PASS."""
        root = _parse_xml(self._braking_xosc("4.0"))
        from src.checks.model_review import check_mr_02
        result = check_mr_02(root, config)
        assert result.status == "PASS", f"Expected PASS for 4.0 m/s², got {result.status}: {result.comment}"

    def test_mr_02_fail_wrong_decel(self, config):
        """GVT decel = 2.0 m/s² (wrong) -> FAIL."""
        root = _parse_xml(self._braking_xosc("2.0"))
        from src.checks.model_review import check_mr_02
        result = check_mr_02(root, config)
        assert result.status == "FAIL", f"Expected FAIL for 2.0 m/s², got {result.status}: {result.comment}"

    def test_mr_02_manual_review_parameterized_decel(self, config):
        """Decel value is $GVT_deceleration param that resolves to 2.0 -> FAIL (resolved from ParameterDeclaration)."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRb_vectorgrp_style"/>
          <ParameterDeclarations>
            <ParameterDeclaration name="GVT_deceleration" parameterType="double" value="2"/>
          </ParameterDeclarations>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
            <ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>
          </Entities>
          <Storyboard>
            <Story name="GVT_Braking">
              <Act name="Act">
                <ManeuverGroup name="MG" maximumExecutionCount="1">
                  <Actors selectTriggeringEntities="false">
                    <EntityRef entityRef="GVT"/>
                  </Actors>
                  <Maneuver name="M">
                    <Event name="E" priority="overwrite">
                      <Action name="A">
                        <PrivateAction>
                          <LongitudinalAction>
                            <SpeedAction>
                              <SpeedActionDynamics dynamicsDimension="rate" dynamicsShape="linear" value="$GVT_deceleration"/>
                              <SpeedActionTarget><AbsoluteTargetSpeed value="0"/></SpeedActionTarget>
                            </SpeedAction>
                          </LongitudinalAction>
                        </PrivateAction>
                      </Action>
                      <StartTrigger/>
                    </Event>
                  </Maneuver>
                </ManeuverGroup>
                <StartTrigger/>
              </Act>
            </Story>
            <StopTrigger/>
          </Storyboard>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.model_review import check_mr_02
        result = check_mr_02(root, config)
        # $GVT_deceleration resolves to 2 m/s² from ParameterDeclaration -> FAIL (not 4.0)
        assert result.status == "FAIL", (
            f"vectorgrp CCR base scenario uses GVT_deceleration=2 m/s² (CCRs default), "
            f"expected FAIL (protocol requires 4.0). Got {result.status}: {result.comment}"
        )

    def test_mr_02_pass_parameterized_correct_decel(self, config):
        """$GVT_deceleration resolves to 4.0 via ParameterDeclaration -> PASS."""
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRb_CorrectDecel"/>
          <ParameterDeclarations>
            <ParameterDeclaration name="GVT_deceleration" parameterType="double" value="4.0"/>
          </ParameterDeclarations>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
            <ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>
          </Entities>
          <Storyboard>
            <Story name="GVT_Braking">
              <Act name="Act">
                <ManeuverGroup name="MG" maximumExecutionCount="1">
                  <Actors selectTriggeringEntities="false">
                    <EntityRef entityRef="GVT"/>
                  </Actors>
                  <Maneuver name="M">
                    <Event name="E" priority="overwrite">
                      <Action name="A">
                        <PrivateAction>
                          <LongitudinalAction>
                            <SpeedAction>
                              <SpeedActionDynamics dynamicsDimension="rate" dynamicsShape="linear" value="$GVT_deceleration"/>
                              <SpeedActionTarget><AbsoluteTargetSpeed value="0"/></SpeedActionTarget>
                            </SpeedAction>
                          </LongitudinalAction>
                        </PrivateAction>
                      </Action>
                      <StartTrigger/>
                    </Event>
                  </Maneuver>
                </ManeuverGroup>
                <StartTrigger/>
              </Act>
            </Story>
            <StopTrigger/>
          </Storyboard>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.model_review import check_mr_02
        result = check_mr_02(root, config)
        assert result.status == "PASS", f"Expected PASS when $GVT_deceleration=4.0, got {result.status}: {result.comment}"


class TestModelReviewSpeedSanity:
    """CH_MR_01 - garbage/incorrect speed values for VUT and Asset."""

    def _xosc_with_speeds(self, vut_speed: str, target_speed: str) -> bytes:
        return f"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRs_Test"/>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
            <ScenarioObject name="GVT"><Vehicle name="GVT"/></ScenarioObject>
          </Entities>
          <Init>
            <Actions>
              <Private entityRef="Ego">
                <PrivateAction><LongitudinalAction><SpeedAction><SpeedActionTarget>
                  <AbsoluteTargetSpeed value="{vut_speed}"/>
                </SpeedActionTarget></SpeedAction></LongitudinalAction></PrivateAction>
              </Private>
              <Private entityRef="GVT">
                <PrivateAction><LongitudinalAction><SpeedAction><SpeedActionTarget>
                  <AbsoluteTargetSpeed value="{target_speed}"/>
                </SpeedActionTarget></SpeedAction></LongitudinalAction></PrivateAction>
              </Private>
            </Actions>
          </Init>
        </OpenSCENARIO>""".encode()

    def test_mr_01_pass_sane_speeds(self, config):
        root = _parse_xml(self._xosc_with_speeds("13.89", "0"))
        from src.checks.model_review import check_mr_01
        result = check_mr_01(root, config)
        assert result.check_id == "CH_MR_01"
        assert result.status == "PASS", f"got {result.status}: {result.comment}"

    def test_mr_01_fail_negative_speed(self, config):
        root = _parse_xml(self._xosc_with_speeds("-5", "0"))
        from src.checks.model_review import check_mr_01
        result = check_mr_01(root, config)
        assert result.status == "FAIL"
        assert "negative" in result.comment.lower()

    def test_mr_01_fail_absurd_speed(self, config):
        """600 km/h = 166.7 m/s is above the 300 km/h sanity bound -> FAIL."""
        root = _parse_xml(self._xosc_with_speeds("166.7", "0"))
        from src.checks.model_review import check_mr_01
        result = check_mr_01(root, config)
        assert result.status == "FAIL"

    def test_mr_01_na_no_speeds(self, config):
        xml = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRs_Test"/>
          <Entities>
            <ScenarioObject name="Ego"><Vehicle name="Ego"/></ScenarioObject>
          </Entities>
        </OpenSCENARIO>"""
        root = _parse_xml(xml)
        from src.checks.model_review import check_mr_01
        result = check_mr_01(root, config)
        assert result.status == "NA"


class TestFunctionalBlock:
    """CH_FB_01 - ENCAP functional / Test-Automation workbook provisioning."""

    def test_fb_01_present_parseable_manual_review(self, config, tmp_path):
        base = "AEB_CCRs_50VUT_0GVT_50Imp"
        (tmp_path / f"{base}.rrscene").write_text("rrscene", encoding="utf-8")
        (tmp_path / config.functional_file_name(base)).write_bytes(_workbook_bytes())
        from src.checks.functional_block import check_fb_01
        result = check_fb_01(tmp_path, config)
        assert result.check_id == "CH_FB_01"
        assert result.category == "FunctionalBlock"
        assert result.status == "MANUAL_REVIEW"

    def test_fb_01_empty_file_fails(self, config, tmp_path):
        base = "AEB_CCRs_50VUT_0GVT_50Imp"
        (tmp_path / f"{base}.rrscene").write_text("rrscene", encoding="utf-8")
        (tmp_path / config.functional_file_name(base)).touch()  # empty -> not a valid zip
        from src.checks.functional_block import check_fb_01
        result = check_fb_01(tmp_path, config)
        assert result.check_id == "CH_FB_01"
        assert result.category == "FunctionalBlock"
        assert result.status == "FAIL"

    def test_fb_01_missing_fails(self, config, tmp_path):
        from src.checks.functional_block import check_fb_01
        result = check_fb_01(tmp_path, config)
        assert result.check_id == "CH_FB_01"
        assert result.status == "FAIL"


class TestGeometry:
    def test_50pct_overlap_static(self):
        from src.geometry import VehicleState, lateral_offset_percentage
        vut    = VehicleState(x=0, y=0,   heading_deg=0, length=4.5, width=1.8, speed_ms=0)
        target = VehicleState(x=0, y=0.9, heading_deg=0, length=4.5, width=1.8, speed_ms=0)
        pct = lateral_offset_percentage(vut, target)
        assert 40 <= pct <= 60, f"Expected ~50%, got {pct:.1f}%"

    def test_100pct_overlap(self):
        from src.geometry import VehicleState, lateral_offset_percentage
        vut    = VehicleState(x=0, y=0, heading_deg=0, length=4.5, width=1.8, speed_ms=0)
        target = VehicleState(x=0, y=0, heading_deg=0, length=4.5, width=1.8, speed_ms=0)
        pct = lateral_offset_percentage(vut, target)
        assert pct >= 95.0, f"Expected ~100%, got {pct:.1f}%"

    def test_zero_overlap(self):
        from src.geometry import VehicleState, lateral_offset_percentage
        vut    = VehicleState(x=0, y=0,  heading_deg=0, length=4.5, width=1.8, speed_ms=0)
        target = VehicleState(x=0, y=20, heading_deg=0, length=4.5, width=1.8, speed_ms=0)
        pct = lateral_offset_percentage(vut, target)
        assert pct < 1.0, f"Expected ~0%, got {pct:.1f}%"

    def test_longitudinal_impact_projection(self):
        from src.geometry import VehicleState, compute_impact_percentage
        vut    = VehicleState(x=-30, y=0, heading_deg=0, length=4.5, width=1.8, speed_ms=14.0)
        target = VehicleState(x=0,   y=0, heading_deg=0, length=4.5, width=1.8, speed_ms=0.0)
        pct = compute_impact_percentage(vut, target, scenario_type="longitudinal")
        assert pct > 10, f"Expected meaningful overlap for head-on longitudinal, got {pct:.1f}%"


# ============================================================
# Config tests (no external files needed)
# ============================================================

class TestConfig:
    def test_config_loads(self, config):
        assert config.lane_width_m == 3.5
        assert config.junction_radius_m == 8.0
        assert config.simulation_time_min_s == 100
        assert config.simulation_time_max_s == 150

    def test_scenario_protocol_lookup(self, config):
        proto = config.scenario_protocol("CCRs")
        assert proto is not None
        assert proto.type == "longitudinal"

    def test_vut_dims(self, config):
        dims = config.vut_dims()
        assert dims.length > 0
        assert dims.width > 0


# ============================================================
# Full pipeline tests (no external files needed)
# ============================================================

class TestFullPipeline:
    def test_run_validation_synthetic(self, tmp_path):
        """
        Build a minimal synthetic scenario directory and run the full validation pipeline.
        Verifies the pipeline runs without crashing and returns structured results.
        """
        xodr = b"""<?xml version="1.0"?>
        <OpenDRIVE>
          <road id="1" length="200" junction="-1">
            <link><successor elementType="road" elementId="2" contactPoint="start"/></link>
            <planView><geometry x="0" y="0" hdg="0" length="200"><line/></geometry></planView>
            <lanes><laneSection s="0"><right>
              <lane id="-1" type="driving">
                <width sOffset="0" a="3.5" b="0" c="0" d="0"/>
                <roadMark type="solid" weight="standard"/>
              </lane>
            </right></laneSection></lanes>
          </road>
          <road id="2" length="200" junction="-1">
            <link><predecessor elementType="road" elementId="1" contactPoint="end"/></link>
            <planView><geometry x="200" y="0" hdg="0" length="200"><line/></geometry></planView>
            <lanes><laneSection s="0"><right>
              <lane id="-1" type="driving">
                <width sOffset="0" a="3.5" b="0" c="0" d="0"/>
                <roadMark type="solid" weight="standard"/>
              </lane>
            </right></laneSection></lanes>
          </road>
        </OpenDRIVE>"""

        xosc = b"""<?xml version="1.0" encoding="UTF-8"?>
        <OpenSCENARIO>
          <FileHeader description="CCRs_Test" author="Test"/>
          <ParameterDeclarations>
            <ParameterDeclaration name="Ego_speed_kph" parameterType="double" value="50"/>
          </ParameterDeclarations>
          <RoadNetwork>
            <LogicFile filepath="CCRs.xodr"/>
          </RoadNetwork>
          <Entities>
            <ScenarioObject name="Ego">
              <Vehicle name="Ego" vehicleCategory="car">
                <Properties><Property name="filepath" value="NCAP/Assets/Car.fbx"/></Properties>
              </Vehicle>
            </ScenarioObject>
            <ScenarioObject name="Target">
              <Vehicle name="Target" vehicleCategory="car">
                <Properties><Property name="filepath" value="NCAP/Assets/GVT.fbx"/></Properties>
              </Vehicle>
            </ScenarioObject>
          </Entities>
          <Storyboard>
            <Init>
              <Actions>
                <Private entityRef="Ego">
                  <PrivateAction>
                    <TeleportAction>
                      <Position><LanePosition roadId="1" laneId="-1" s="10" offset="0"/></Position>
                    </TeleportAction>
                  </PrivateAction>
                  <PrivateAction>
                    <LongitudinalAction>
                      <SpeedAction><SpeedActionTarget>
                        <AbsoluteTargetSpeed value="13.89"/>
                      </SpeedActionTarget></SpeedAction>
                    </LongitudinalAction>
                  </PrivateAction>
                </Private>
                <Private entityRef="Target">
                  <PrivateAction>
                    <TeleportAction>
                      <Position><WorldPosition x="100" y="0" z="0" h="0"/></Position>
                    </TeleportAction>
                  </PrivateAction>
                  <PrivateAction>
                    <LongitudinalAction>
                      <SpeedAction><SpeedActionTarget>
                        <AbsoluteTargetSpeed value="0"/>
                      </SpeedActionTarget></SpeedAction>
                    </LongitudinalAction>
                  </PrivateAction>
                </Private>
              </Actions>
            </Init>
            <Story name="Story">
              <Act name="Act">
                <ManeuverGroup name="MG" maximumExecutionCount="1">
                  <Actors selectTriggeringEntities="false">
                    <EntityRef entityRef="Ego"/>
                  </Actors>
                  <Maneuver name="M">
                    <Event name="E" priority="overwrite">
                      <Action name="A">
                        <PrivateAction>
                          <RoutingAction>
                            <FollowTrajectoryAction>
                              <TimeReference><Timing domainAbsoluteRelative="relative" scale="1" offset="0"/></TimeReference>
                              <Trajectory name="T" closed="false">
                                <Shape><Polyline>
                                  <Vertex><Position><WorldPosition x="10" y="0" z="0" h="0"/></Position></Vertex>
                                  <Vertex><Position><WorldPosition x="200" y="0" z="0" h="0"/></Position></Vertex>
                                </Polyline></Shape>
                              </Trajectory>
                            </FollowTrajectoryAction>
                          </RoutingAction>
                        </PrivateAction>
                      </Action>
                      <StartTrigger/>
                    </Event>
                  </Maneuver>
                </ManeuverGroup>
                <StartTrigger/>
              </Act>
            </Story>
            <StopTrigger>
              <ConditionGroup>
                <Condition name="Stop" delay="0" conditionEdge="none">
                  <ByValueCondition>
                    <SimulationTimeCondition value="120" rule="greaterThan"/>
                  </ByValueCondition>
                </Condition>
              </ConditionGroup>
            </StopTrigger>
          </Storyboard>
        </OpenSCENARIO>"""

        scenario_dir = tmp_path / "CCRs_test"
        scenario_dir.mkdir()
        (scenario_dir / "CCRs.xodr").write_bytes(xodr)
        (scenario_dir / "CCRs.xosc").write_bytes(xosc)
        for ext in ("rrscene", "rrscenario", "rd", "xml", "txt"):
            (scenario_dir / f"CCRs.{ext}").touch()
        (scenario_dir / "TA.xml").write_text("<TA/>", encoding="utf-8")

        from validator import run_validation
        results, stats = run_validation(scenario_dir, skip_rd=True)

        assert len(results) > 0
        assert stats.total > 0
        assert isinstance(stats.pass_rate, float)
        parse_errors = {r.check_id: r.comment for r in results
                        if r.status == "FAIL" and "error" in r.comment.lower()}
        assert not parse_errors, f"Unexpected parse errors: {parse_errors}"

    def test_excel_report_created(self, tmp_path, config):
        from src.models import CheckResult, SummaryStats
        from src.reporter import write_excel

        results = [
            CheckResult(check_id="CH_NM_01", category="Naming",
                        description="Test", status="PASS", comment="ok"),
            CheckResult(check_id="CH_NM_02", category="Naming",
                        description="Test", status="FAIL", comment="mismatch"),
            CheckResult(check_id="CH_RD_01", category="Road",
                        description="Test", status="NA", comment=""),
        ]
        stats = SummaryStats(
            scenario_name="Test_CCRs", run_timestamp="2026-05-25 22:00:00",
            protocol_version="AEB_CC_4.3.1", total=3, passed=1, failed=1,
            manual=0, na=1, pass_rate=33.3, critical_failures=["CH_NM_02"],
        )
        out = tmp_path / "test_report.xlsx"
        write_excel(results, stats, out)
        assert out.exists()
        assert out.stat().st_size > 0
        import openpyxl
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames[:3] == ["Validation", "Issues Log", "Run Summary"]
        assert [wb["Validation"].cell(row=3, column=i).value for i in range(1, 10)] == [
            "Check ID",
            "Category",
            "Check name",
            "Result",
            "Comment",
            "Source file",
            "Severity",
            "Automatable or Manual",
            "Timestamp",
        ]

    def test_csv_report_created(self, tmp_path):
        from src.models import CheckResult, SummaryStats
        from src.reporter import write_csv

        results = [
            CheckResult(check_id="CH_NM_01", category="Naming",
                        description="Test", status="PASS", comment="ok"),
        ]
        stats = SummaryStats(
            scenario_name="Test", run_timestamp="2026-05-25 22:00:00",
            protocol_version="test", total=1, passed=1, failed=0,
            manual=0, na=0, pass_rate=100.0, critical_failures=[],
        )
        out = tmp_path / "test_report.csv"
        write_csv(results, stats, out)
        assert out.exists()
        content = out.read_text()
        assert "CH_NM_01" in content
        assert "Yes" in content
        assert "=== RUN SUMMARY ===" not in content
