from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
CONFIG = ROOT / "config.json"


def _xodr(lane_width: str = "3.5") -> bytes:
    return f"""<?xml version="1.0"?>
<OpenDRIVE>
  <road id="1" length="200" junction="-1">
    <link><successor elementType="road" elementId="2" contactPoint="start"/></link>
    <planView><geometry x="0" y="0" hdg="0" length="200"><line/></geometry></planView>
    <lanes><laneSection s="0"><right>
      <lane id="-1" type="driving">
        <width sOffset="0" a="{lane_width}" b="0" c="0" d="0"/>
        <roadMark type="solid"/>
      </lane>
    </right></laneSection></lanes>
  </road>
  <road id="2" length="200" junction="-1">
    <link><predecessor elementType="road" elementId="1" contactPoint="end"/></link>
    <planView><geometry x="200" y="0" hdg="0" length="200"><line/></geometry></planView>
    <lanes><laneSection s="0"><right>
      <lane id="-1" type="driving">
        <width sOffset="0" a="{lane_width}" b="0" c="0" d="0"/>
        <roadMark type="solid"/>
      </lane>
    </right></laneSection></lanes>
  </road>
</OpenDRIVE>""".encode()


def _xosc(vut_speed_ms: str = "13.89", sim_time: str = "50") -> bytes:
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<OpenSCENARIO>
  <FileHeader description="CCRs_70kph" author="Synthetic"/>
  <ParameterDeclarations>
    <ParameterDeclaration name="Ego_speed_kph" parameterType="double" value="50"/>
  </ParameterDeclarations>
  <Entities>
    <ScenarioObject name="Ego">
      <Vehicle name="Ego" vehicleCategory="car">
        <Properties><Property name="filepath" value="OEM/Ego.fbx"/></Properties>
      </Vehicle>
    </ScenarioObject>
    <ScenarioObject name="GVT">
      <Vehicle name="GVT" vehicleCategory="car">
        <Properties><Property name="filepath" value="NCAP/Asset/GVT.fbx"/></Properties>
      </Vehicle>
    </ScenarioObject>
  </Entities>
  <Storyboard>
    <Init><Actions>
      <Private entityRef="Ego">
        <PrivateAction><TeleportAction><Position>
          <WorldPosition x="-30" y="0" z="0" h="0"/>
        </Position></TeleportAction></PrivateAction>
        <PrivateAction><LongitudinalAction><SpeedAction><SpeedActionTarget>
          <AbsoluteTargetSpeed value="{vut_speed_ms}"/>
        </SpeedActionTarget></SpeedAction></LongitudinalAction></PrivateAction>
      </Private>
      <Private entityRef="GVT">
        <PrivateAction><TeleportAction><Position>
          <WorldPosition x="0" y="0" z="0" h="0"/>
        </Position></TeleportAction></PrivateAction>
        <PrivateAction><LongitudinalAction><SpeedAction><SpeedActionTarget>
          <AbsoluteTargetSpeed value="0"/>
        </SpeedActionTarget></SpeedAction></LongitudinalAction></PrivateAction>
      </Private>
    </Actions></Init>
    <Story name="Story"><Act name="Act">
      <ManeuverGroup name="MG" maximumExecutionCount="1">
        <Actors selectTriggeringEntities="false"><EntityRef entityRef="Ego"/></Actors>
        <Maneuver name="M"><Event name="E" priority="overwrite">
          <Action name="A"><PrivateAction><RoutingAction><FollowTrajectoryAction>
            <TimeReference><Timing domainAbsoluteRelative="relative" scale="1" offset="0"/></TimeReference>
            <Trajectory name="T" closed="false"><Shape><Polyline>
              <Vertex><Position><WorldPosition x="-30" y="0" z="0" h="0"/></Position></Vertex>
              <Vertex><Position><WorldPosition x="20" y="0" z="0" h="0"/></Position></Vertex>
            </Polyline></Shape></Trajectory>
          </FollowTrajectoryAction></RoutingAction></PrivateAction></Action>
          <StartTrigger><ConditionGroup><Condition name="speed" delay="0" conditionEdge="rising">
            <ByEntityCondition><TriggeringEntities triggeringEntitiesRule="any">
              <EntityRef entityRef="Ego"/>
            </TriggeringEntities><EntityCondition><SpeedCondition value="13" rule="greaterThan"/></EntityCondition></ByEntityCondition>
          </Condition></ConditionGroup></StartTrigger>
        </Event></Maneuver>
      </ManeuverGroup>
      <StartTrigger/>
    </Act></Story>
    <StopTrigger><ConditionGroup><Condition name="stop" delay="0" conditionEdge="none">
      <ByValueCondition><SimulationTimeCondition value="{sim_time}" rule="greaterThan"/></ByValueCondition>
    </Condition></ConditionGroup></StopTrigger>
  </Storyboard>
</OpenSCENARIO>""".encode()


def _rd() -> bytes:
    return b"""<?xml version="1.0"?>
<Routes>
  <Route name="Ego"><Road id="1"/><Road id="2"/></Route>
  <Route name="GVT"><Road id="1"/><Road id="2"/></Route>
</Routes>"""


def _workbook_bytes() -> bytes:
    """Minimal valid OOXML (zip) so CH_FB_01's zipfile check treats it as a workbook."""
    import io
    import zipfile

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", "<Types/>")
    return buf.getvalue()


def create_scenario(
    base_dir: Path,
    name: str = "AEB_CCRs_50VUT_0GVT_50Imp",
    *,
    lane_width: str = "3.50",
    vut_speed_ms: str = "13.89",
    sim_time: str = "50",
    include_rd: bool = True,
    include_ta: bool = True,
    base_override: dict[str, str] | None = None,
) -> Path:
    scenario_dir = base_dir / name
    scenario_dir.mkdir()
    base_override = base_override or {}

    base = base_override.get("default", name)
    (scenario_dir / f"{base_override.get('rrscene', base)}.rrscene").write_text("rrscene", encoding="utf-8")
    (scenario_dir / f"{base_override.get('rrscenario', base)}.rrscenario").write_text("rrscenario", encoding="utf-8")
    (scenario_dir / f"{base_override.get('xosc', base)}.xosc").write_bytes(_xosc(vut_speed_ms, sim_time))
    (scenario_dir / f"{base_override.get('xodr', base)}.xodr").write_bytes(_xodr(lane_width))
    if include_rd:
        (scenario_dir / f"{base_override.get('rd', base)}.rd").write_bytes(_rd())
    (scenario_dir / f"{base_override.get('xml', base)}.xml").write_text("<ScenarioMeta/>", encoding="utf-8")
    (scenario_dir / f"{base_override.get('txt', base)}.txt").write_text("notes", encoding="utf-8")
    # Macro workbook is always required; the ENCAP functional workbook is toggled by include_ta.
    (scenario_dir / f"MACRO_{base}.xlsx").write_bytes(_workbook_bytes())
    if include_ta:
        (scenario_dir / f"ENCAP_Scenario_func_{base}.xlsm").write_bytes(_workbook_bytes())
    return scenario_dir


def run_cli(scenario_dir: Path, output_dir: Path, *args: str) -> subprocess.CompletedProcess[str]:
    cmd = [
        sys.executable,
        str(ROOT / "validator.py"),
        str(scenario_dir),
        "--config",
        str(CONFIG),
        "--output",
        str(output_dir),
        *args,
    ]
    return subprocess.run(cmd, cwd=ROOT, text=True, capture_output=True)


def latest_xlsx(output_dir: Path) -> Path:
    return sorted(output_dir.glob("*.xlsx"))[-1]


def validation_rows(workbook: Path) -> dict[str, str]:
    ws = openpyxl.load_workbook(workbook)["Validation"]
    rows: dict[str, str] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0]:
            rows[str(row[0])] = str(row[3])
    return rows


def test_cli_success_and_report_contract(tmp_path: Path) -> None:
    scenario_dir = create_scenario(tmp_path)
    out = tmp_path / "out"
    result = run_cli(scenario_dir, out)
    assert result.returncode == 0, result.stdout + result.stderr
    assert (out / "validation_run.log").exists()
    assert "Exit status: 0" in (out / "validation_run.log").read_text(encoding="utf-8")

    wb = openpyxl.load_workbook(latest_xlsx(out))
    assert wb.sheetnames[:3] == ["Validation", "Issues Log", "Run Summary"]
    assert [wb["Validation"].cell(row=3, column=i).value for i in range(1, 10)] == [
        "Check ID",
        "Category",
        "Check name",
        "Result",
        "Comment",
        "Source file",
        "Severity",
        "Automatable or Manual",
        "Timestamp",
    ]
    rows = validation_rows(latest_xlsx(out))
    assert rows["CH_NM_03"] == "Yes"
    assert rows["CH_MD_02"] == "Yes"


def test_cli_failure_exit_code_and_issues_log(tmp_path: Path) -> None:
    scenario_dir = create_scenario(tmp_path, include_ta=False)
    out = tmp_path / "out"
    result = run_cli(scenario_dir, out)
    assert result.returncode == 1
    wb = openpyxl.load_workbook(latest_xlsx(out))
    issue_values = [cell.value for cell in wb["Issues Log"]["A"]]
    assert "CH_FB_01" in issue_values
    assert validation_rows(latest_xlsx(out))["CH_FB_01"] == "No"


def test_no_rd_mode_skips_cleanly(tmp_path: Path) -> None:
    scenario_dir = create_scenario(tmp_path, include_rd=False)
    out = tmp_path / "out"
    result = run_cli(scenario_dir, out, "--no-rd")
    assert result.returncode == 0, result.stdout + result.stderr
    rows = validation_rows(latest_xlsx(out))
    # .rd handling folded into CH_NM_03: all other required files present -> Yes under --no-rd.
    assert rows["CH_NM_03"] == "Yes"
    assert rows["CH_MD_01"] == "NA"


def test_missing_rd_without_no_rd_fails(tmp_path: Path) -> None:
    scenario_dir = create_scenario(tmp_path, include_rd=False)
    out = tmp_path / "out"
    result = run_cli(scenario_dir, out)
    assert result.returncode == 1
    rows = validation_rows(latest_xlsx(out))
    # missing .rd is reported by CH_NM_03 (old CH_NM_07 folded in)
    assert rows["CH_NM_03"] == "No"
    assert rows["CH_MD_02"] == "No"


def test_bad_lane_width_and_bad_speed_fail(tmp_path: Path) -> None:
    bad_lane = create_scenario(tmp_path, name="invalid_bad_lane_width", lane_width="3.0")
    out_lane = tmp_path / "out_lane"
    result_lane = run_cli(bad_lane, out_lane)
    assert result_lane.returncode == 1
    assert validation_rows(latest_xlsx(out_lane))["CH_RD_01"] == "No"

    bad_speed = create_scenario(tmp_path, name="invalid_bad_speed", vut_speed_ms="55")
    out_speed = tmp_path / "out_speed"
    result_speed = run_cli(bad_speed, out_speed)
    assert result_speed.returncode == 1
    rows = validation_rows(latest_xlsx(out_speed))
    assert rows["CH_SC_18"] == "No"


def test_base_name_mismatch_detected(tmp_path: Path) -> None:
    scenario_dir = create_scenario(tmp_path, base_override={"xosc": "Different_CCRs"})
    out = tmp_path / "out"
    result = run_cli(scenario_dir, out)
    assert result.returncode == 1
    rows = validation_rows(latest_xlsx(out))
    # base-name mismatch + duplicate/case detection folded into CH_NM_04
    assert rows["CH_NM_04"] == "No"


def test_quiet_suppresses_console_but_writes_log(tmp_path: Path) -> None:
    scenario_dir = create_scenario(tmp_path)
    out = tmp_path / "out"
    result = run_cli(scenario_dir, out, "--quiet")
    assert result.returncode == 0
    assert result.stdout == ""
    assert result.stderr == ""
    assert "CLI command:" in (out / "validation_run.log").read_text(encoding="utf-8")


def test_invalid_config_path_fails_early(tmp_path: Path) -> None:
    scenario_dir = create_scenario(tmp_path)
    cmd = [
        sys.executable,
        str(ROOT / "validator.py"),
        str(scenario_dir),
        "--config",
        str(tmp_path / "missing.json"),
        "--output",
        str(tmp_path / "out"),
    ]
    result = subprocess.run(cmd, cwd=ROOT, text=True, capture_output=True)
    assert result.returncode == 1
    assert "does not exist" in result.stderr


def test_security_payloads_are_not_expanded(tmp_path: Path) -> None:
    from src.parsers import xosc
    from src.checks.functional_block import check_fb_01
    from src.models import Config

    xxe = b"""<?xml version="1.0"?>
<!DOCTYPE data [<!ENTITY xxe SYSTEM "file:///etc/passwd">]>
<OpenSCENARIO><FileHeader description="&xxe;"/></OpenSCENARIO>"""
    path = tmp_path / "payload.xosc"
    path.write_bytes(xxe)
    try:
        root = xosc.load(path)
        assert "root:" not in "".join(root.xpath("//FileHeader/@description"))
    except Exception:
        pass

    # CH_FB_01 must never XML-parse the functional workbook (it is a zip/OOXML), so an
    # XXE payload disguised as the functional file is rejected as "not a valid workbook",
    # never expanded.
    func_dir = tmp_path / "func"
    func_dir.mkdir()
    (func_dir / "func.rrscene").write_text("rrscene", encoding="utf-8")
    (func_dir / "ENCAP_Scenario_func_func.xlsm").write_bytes(xxe.replace(b"OpenSCENARIO", b"TA"))
    result = check_fb_01(func_dir, Config.load(CONFIG))
    assert result.status in {"FAIL", "MANUAL_REVIEW"}


def test_real_examples_smoke_validation(tmp_path: Path) -> None:
    examples = ROOT / "examples"
    if not examples.exists():
        pytest.skip("real examples directory is not present")
    for scenario_dir in sorted(p for p in examples.iterdir() if p.is_dir()):
        out = tmp_path / scenario_dir.name
        result = run_cli(scenario_dir, out, "--no-rd")
        assert result.returncode == 1
        assert latest_xlsx(out).exists()
        rows = validation_rows(latest_xlsx(out))
        assert rows["CH_NM_03"] == "No"
        assert rows["CH_FB_01"] == "No"
