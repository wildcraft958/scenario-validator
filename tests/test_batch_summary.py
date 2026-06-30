"""Tests for the root-level batch summary workbook (reporter.write_batch_summary)."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from src.reporter import _SUMMARY_HEADERS, write_batch_summary
from src.rollup import BatchSummaryMeta, ScenarioRow

_GREEN = "FF92D050"
_RED = "FFFF0000"
_GREY = "FFD3D3D3"


def _rows() -> list[ScenarioRow]:
    return [
        ScenarioRow(batch="Batch 1", category="Car_to_Car", scenario="AEB_clean",
                    total=43, automated=43, passed=43, failed=0, manual=0, na=0,
                    confidence="High", verdict="P", advice="All automated checks passed",
                    path="/data/Batch 1/Car_to_Car/AEB_clean"),
        ScenarioRow(batch="Batch 1", category="Car_to_Car", scenario="AEB_fail",
                    total=43, automated=40, passed=30, failed=2, manual=3, na=8,
                    confidence="Medium", verdict="R", advice="Verify 2 failed: CH_SC_07; 3 manual to review"),
        ScenarioRow(batch="Batch 1", category="Car_to_Bike", scenario="AEB_crash",
                    verdict="ERROR", advice="RUN ERROR: ValueError: boom - not validated, re-run"),
    ]


def _meta() -> BatchSummaryMeta:
    return BatchSummaryMeta(
        root="/x/Batch 1", run_timestamp="2026-06-29 10:00:00",
        discovered=3, validated=2, skipped=1, errored=1, checks_per_scenario=43,
        incompatible_dirs=["RR Scenarios/foo"],
        error_details=[("01_Car_to_Bike/AEB_crash", "ValueError: boom")],
        report_warnings=[("01_Car_to_Car/AEB_fail", "PermissionError: file open in Excel")],
    )


def _find_header_row(ws) -> int:
    for r in range(1, 40):
        if ws.cell(row=r, column=1).value == "S/No":
            return r
    raise AssertionError("header row not found")


def test_sheets_and_headers(tmp_path: Path):
    out = tmp_path / "Summary.xlsx"
    write_batch_summary(_rows(), _meta(), out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Summary", "Skipped & Errors"]
    ws = wb["Summary"]
    hr = _find_header_row(ws)
    headers = [ws.cell(row=hr, column=c).value for c in range(1, len(_SUMMARY_HEADERS) + 1)]
    assert headers == _SUMMARY_HEADERS


def test_one_row_per_scenario(tmp_path: Path):
    out = tmp_path / "Summary.xlsx"
    write_batch_summary(_rows(), _meta(), out)
    ws = openpyxl.load_workbook(out)["Summary"]
    hr = _find_header_row(ws)
    scenarios = [ws.cell(row=hr + 1 + i, column=4).value for i in range(3)]
    assert scenarios == ["AEB_clean", "AEB_fail", "AEB_crash"]
    # S/No is 1..n
    assert [ws.cell(row=hr + 1 + i, column=1).value for i in range(3)] == [1, 2, 3]


def test_path_column_holds_absolute_path(tmp_path: Path):
    out = tmp_path / "Summary.xlsx"
    write_batch_summary(_rows(), _meta(), out)
    ws = openpyxl.load_workbook(out)["Summary"]
    hr = _find_header_row(ws)
    pcol = len(_SUMMARY_HEADERS)  # Path is the last column
    assert ws.cell(row=hr, column=pcol).value == "Path"
    assert ws.cell(row=hr + 1, column=pcol).value == "/data/Batch 1/Car_to_Car/AEB_clean"


def test_top_box_values(tmp_path: Path):
    out = tmp_path / "Summary.xlsx"
    write_batch_summary(_rows(), _meta(), out)
    ws = openpyxl.load_workbook(out)["Summary"]
    box = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=4).value for r in range(2, 12)}
    assert box["Root folder"] == "/x/Batch 1"
    assert box["Scenarios discovered"] == 3
    assert box["Pass (P)"] == 1
    assert box["Review (R)"] == 1
    assert box["Errored"] == 1


def test_verdict_cell_colours(tmp_path: Path):
    out = tmp_path / "Summary.xlsx"
    write_batch_summary(_rows(), _meta(), out)
    ws = openpyxl.load_workbook(out)["Summary"]
    hr = _find_header_row(ws)
    assert ws.cell(row=hr + 1, column=12).fill.fgColor.rgb == _GREEN   # P
    assert ws.cell(row=hr + 2, column=12).fill.fgColor.rgb == _RED     # R with failure
    assert ws.cell(row=hr + 3, column=12).fill.fgColor.rgb == _GREY    # ERROR


def test_skipped_and_errors_sheet_lists_everything(tmp_path: Path):
    out = tmp_path / "Summary.xlsx"
    write_batch_summary(_rows(), _meta(), out)
    ws = openpyxl.load_workbook(out)["Skipped & Errors"]
    text = "\n".join(
        str(ws.cell(row=r, column=c).value)
        for r in range(1, ws.max_row + 1) for c in (1, 2)
        if ws.cell(row=r, column=c).value is not None
    )
    assert "RR Scenarios/foo" in text
    assert "01_Car_to_Bike/AEB_crash" in text
    assert "ValueError: boom" in text
    # a failed report write (validation still succeeded) is surfaced, not hidden
    assert "01_Car_to_Car/AEB_fail" in text
    assert "PermissionError: file open in Excel" in text


def test_empty_run_does_not_crash(tmp_path: Path):
    out = tmp_path / "Summary.xlsx"
    meta = BatchSummaryMeta(root="/x", run_timestamp="t", discovered=0, validated=0,
                            skipped=0, errored=0)
    write_batch_summary([], meta, out)
    assert out.is_file()
