#!/usr/bin/env python3
"""Cross-check the automated validator against the hand-filled _Review.xlsx files.

Each scenario folder may ship a human-completed reviewer checklist
(`<base>_Review.xlsx`, sheet ChecklistFinal, Self Review column). Taking those as
ground truth, this compares our automated verdict per check id and reports:

  * agreement on the directly comparable checks,
  * FALSE PASS on our side  - reviewer said No, we said Yes (the dangerous case),
  * reviewer-miss candidates - we said No where the reviewer said Yes/NA,
  * coverage we add          - we decide a check the reviewer left N/A.

Only checks present in BOTH the run and the checklist are compared (our extra
NM_04-06 and the reviewer-only MD_06-11 / FB_02 are skipped).

Usage:
    python tools/crosscheck_reviews.py Eval_Data [--report PATH]
"""
from __future__ import annotations

import argparse
import glob
import logging
import sys
from collections import Counter
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import openpyxl  # noqa: E402

COMPARABLE = (
    [f"CH_NM_0{i}" for i in (1, 2, 3)]
    + [f"CH_RD_0{i}" for i in range(1, 7)]
    + [f"CH_SC_{i:02d}" for i in range(1, 23)]
    + [f"CH_MD_0{i}" for i in range(1, 6)]
    + ["CH_MR_01", "CH_MR_02", "CH_FB_01"]
)


def _norm(v: object) -> str:
    u = str(v).strip().upper()
    if u in ("YES", "Y"):
        return "Yes"
    if u in ("NO", "N"):
        return "No"
    if u == "NA":
        return "NA"
    return u  # leave anything else (e.g. free text) as-is


def human_verdicts(xlsx: Path) -> dict[str, str]:
    ws = openpyxl.load_workbook(xlsx, data_only=True)["ChecklistFinal"]
    out: dict[str, str] = {}
    for r in range(9, ws.max_row + 1):
        cid = ws.cell(row=r, column=3).value
        self_review = ws.cell(row=r, column=5).value
        review1 = ws.cell(row=r, column=6).value
        if not cid:
            continue
        v = self_review if self_review not in (None, "") else review1
        if v is not None:
            out[str(cid).strip()] = _norm(v)
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Cross-check validator vs hand-filled reviews")
    parser.add_argument("root", help="Root directory to scan for *_Review.xlsx")
    parser.add_argument("--report", default=None, help="Write the markdown summary here")
    args = parser.parse_args()

    logging.disable(logging.CRITICAL)
    from validator import run_validation

    reviews = sorted(glob.glob(str(Path(args.root) / "**" / "*_Review.xlsx"), recursive=True))
    matches = 0
    mismatch_types: Counter = Counter()
    defer: Counter = Counter()
    false_pass: list[tuple[str, str, str]] = []     # reviewer No, ours Yes
    we_fail: list[tuple[str, str, str, str]] = []    # ours No, reviewer Yes/NA
    human_dist: Counter = Counter()

    for rv in reviews:
        sdir = Path(rv).parent
        hv = human_verdicts(Path(rv))
        for v in hv.values():
            human_dist[v] += 1
        skip_rd = not any(p.suffix == ".rd" for p in sdir.iterdir() if p.is_file())
        results, _ = run_validation(sdir, skip_rd=skip_rd)
        ours = {r.check_id: r for r in results}
        for cid in COMPARABLE:
            if cid not in hv or cid not in ours:
                continue
            h = hv[cid]
            o = ours[cid].result
            if o == "Manual":
                defer[h] += 1
                continue
            if o == h:
                matches += 1
            else:
                mismatch_types[f"{h}->{o}"] += 1
                if h == "No" and o == "Yes":
                    false_pass.append((sdir.name, cid, ours[cid].comment[:120]))
                elif o == "No":
                    we_fail.append((sdir.name, cid, h, ours[cid].comment[:120]))

    total = matches + sum(mismatch_types.values())
    lines = ["# Cross-check vs hand-validated reviews", ""]
    lines.append(f"- Reviews found: **{len(reviews)}**")
    lines.append(f"- Reviewer verdict mix (Yes/NA only are real): {dict(human_dist)}")
    lines.append(f"- Comparable decisions: **{total}**  |  agreement: "
                 f"**{matches / total * 100:.1f}%**" if total else "- no comparable decisions")
    lines.append(f"- Our Manual defers (we measured, asked for confirm): {dict(defer)}")
    lines.append(f"- FALSE PASS (reviewer No, ours Yes): **{len(false_pass)}**")
    lines.append("")
    lines.append("## Mismatch types (reviewer -> ours)")
    for k, v in sorted(mismatch_types.items()):
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## Our FAIL where reviewer passed (reviewer-miss candidates / our strictness)")
    for scn, cid, h, comment in we_fail:
        lines.append(f"- {scn} | {cid} | reviewer={h} | {comment}")
    if false_pass:
        lines.append("")
        lines.append("## FALSE PASS - investigate")
        for scn, cid, comment in false_pass:
            lines.append(f"- {scn} | {cid} | {comment}")

    report = "\n".join(lines)
    print(report)
    if args.report:
        Path(args.report).write_text(report + "\n", encoding="utf-8")
        print(f"\nWritten: {args.report}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
