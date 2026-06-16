#!/usr/bin/env python3
"""Generate config.xlsx from config.json - the non-developer configuration interface.

Usage:
    python tools/make_config_xlsx.py [--json PATH] [--out PATH]

The validator accepts either file directly:
    python validator.py <scenario_dir> --config config.xlsx

Sheets produced (the loader in src/models.py reads exactly this layout):
    Protocol Constants  - EuroNCAP values. DO NOT EDIT unless the protocol changes.
    Site Settings       - entity names, file lists, tolerances. Edit freely.
    Scenarios           - one row per scenario family. Add rows for new scenarios.
    Vehicle Dimensions  - bounding-box fallbacks used when the .xosc has none.
    Curve Radii         - protocol Part 2 turn radii (EuroNCAP Table 1.2.4).
    Sim Time Bands      - speed-dependent simulation-time limits.

Re-run this script after devs change config.json to keep the two formats in sync.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

_ROOT = Path(__file__).resolve().parent.parent

# key → (sheet, plain-language description). Lists are written comma-separated.
# NOTE: `target_type_to_category` (GVT->Vehicle, EPTa->Pedestrian, ...) is intentionally
# NOT exposed here. It is a fixed EuroNCAP protocol map kept as a model default in
# src/models.py (edited in code, not per-site), so it stays out of the Excel sheets and
# the json<->xlsx round-trip self-test below.
_KEY_SHEET = {
    # ---- Protocol Constants (EuroNCAP - do not edit) ----
    "protocol_version": ("Protocol Constants", "Protocol label printed on every report header."),
    "lane_width_m": ("Protocol Constants", "EuroNCAP lane width. CH_RD_01 checks every driving lane against this."),
    "junction_radius_m": ("Protocol Constants", "EuroNCAP junction corner (kerb) radius. CH_RD_03."),
    "simulation_time_min_s": ("Protocol Constants", "Fallback minimum simulation time when VUT speed is unknown. CH_SC_04."),
    "simulation_time_max_s": ("Protocol Constants", "Fallback maximum simulation time. CH_SC_04."),
    "expected_decel_ms2": ("Protocol Constants", "Required braking deceleration for braking scenarios. CH_MR_02."),
    # ---- Site Settings (edit freely) ----
    "traffic_handedness": ("Site Settings", "LHT = drive on left (Japan/India/UK - EuroNCAP default). RHT inverts Farside/Nearside."),
    "vut_entity_names": ("Site Settings", "Names accepted for the vehicle under test (comma-separated)."),
    "encap_actor_names": ("Site Settings", "Allowed target actor name prefixes per EuroNCAP (comma-separated)."),
    "sov_entity_names": ("Site Settings", "Entities exempt from the NCAP-asset-folder rule (SOV may be a real vehicle)."),
    "static_target_name_patterns": ("Site Settings", "Name patterns treated as static obstructions (must have speed 0)."),
    "stationary_target_name_patterns": ("Site Settings", "Name patterns for VRU targets that start stationary."),
    "required_file_extensions": ("Site Settings", "File extensions every scenario folder must contain (CH_NM_03)."),
    "required_standalone_files": ("Site Settings", "Exact filenames that must exist in every scenario folder (rarely used)."),
    "optional_standalone_files": ("Site Settings", "Files reported if present but never required (e.g. catalogs)."),
    "allowed_programs": ("Site Settings", "Programs allowed as the 1st filename token, e.g. AEB (CH_NM_02)."),
    "target_type_tokens": ("Site Settings", "Target-type tokens valid in the 4th filename slot: GVT, EPTa, EPTc, EBTa, EMT (CH_NM_02)."),
    "vut_speed_suffix": ("Site Settings", "Suffix marking the VUT-speed token in the filename (default VUT)."),
    "impact_suffix": ("Site Settings", "Suffix marking the impact-overlap token in the filename (default Imp)."),
    "allowed_impact_overlaps": ("Site Settings", "Impact-overlap %s the filename token may take (CH_NM_02); exact value is checked by SC_16/17."),
    "functional_file_pattern": ("Site Settings", "ENCAP functional/Test-Automation workbook name; {base} = scenario base name (CH_FB_01)."),
    "macro_file_pattern": ("Site Settings", "Macro workbook name; {base} = scenario base name (CH_NM_03)."),
    "review_file_pattern": ("Site Settings", "Manual review checklist name; {base} = scenario base name (optional, being phased out)."),
    "required_associated_roles": ("Site Settings", "Which affix files are mandatory: functional, macro (review is optional)."),
    "junction_intersection_min_spread_deg": ("Site Settings", "A junction counts as an intersection (RD_03-06, SC_10) when its incoming roads differ in heading by more than this (deg). Auto-detected from the .xodr - no scenario list."),
    "extra_scenario_prefixes": ("Site Settings", "Extra valid scenario name prefixes that have no row in the Scenarios sheet."),
    "junction_waypoint_radius_m": ("Site Settings", "How close to a junction a waypoint must be to count as 'at the junction'."),
    # tolerances / tuning
    "lane_width_tolerance_m": ("Site Settings", "Allowed deviation from the 3.5 m lane width."),
    "junction_radius_tolerance_m": ("Site Settings", "Allowed deviation from the 8 m junction radius."),
    "impact_tolerance_pct": ("Site Settings", "Allowed deviation for crossing/turning impact % (CH_SC_16)."),
    "longitudinal_impact_tolerance_pct": ("Site Settings", "Allowed deviation for longitudinal impact % (CH_SC_17)."),
    "decel_tolerance_ms2": ("Site Settings", "Allowed deviation from the -4 m/s² braking value."),
    "speed_sanity_max_kmh": ("Site Settings", "Speeds above this are flagged as garbage data (CH_MR_01)."),
    "east_heading_tolerance_deg": ("Site Settings", "Heading tolerance used by the lane-side check (CH_SC_05)."),
    "cardinal_heading_tolerance_deg": ("Site Settings", "How close to 0/90/180/270° the VUT heading must be (CH_SC_06)."),
    "right_lane_offset_threshold_m": ("Site Settings", "Lateral offset that decides left vs right lane (CH_SC_05)."),
    "curvature_min_heading_delta_rad": ("Site Settings", "Minimum per-vertex heading change to count as turning (CH_SC_07)."),
    "curvature_min_segment_length_m": ("Site Settings", "Minimum segment length used in turn-radius estimation (CH_SC_07)."),
    "curve_radius_tolerance_pct": ("Site Settings", "Allowed deviation from the protocol Part 2 turn radius (CH_SC_07)."),
}

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_LOCK_FILL = PatternFill("solid", fgColor="F2DCDB")


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def build_workbook(raw: dict) -> Workbook:
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    # ---- key/value sheets ----
    sheets = {name: wb.create_sheet(name) for name in ("Protocol Constants", "Site Settings")}
    for ws in sheets.values():
        ws.append(["key", "value", "what it is / when to change it"])
        _style_header(ws, 3)
        _autosize(ws, [34, 40, 95])

    # extra prefixes = valid_prefixes that have no scenarios entry
    scenario_prefixes = {k.split("-")[0] for k in raw.get("scenarios", {})}
    extra_prefixes = [
        p for p in raw.get("naming_convention", {}).get("valid_prefixes", [])
        if p not in scenario_prefixes
    ]

    values = dict(raw)
    values["extra_scenario_prefixes"] = extra_prefixes
    for key, (sheet, desc) in _KEY_SHEET.items():
        value = values.get(key)
        if isinstance(value, list):
            value = ", ".join(str(v) for v in value)
        sheets[sheet].append([key, value, desc])

    ws_const = sheets["Protocol Constants"]
    for row in ws_const.iter_rows(min_row=2):
        for cell in row:
            cell.fill = _LOCK_FILL
    ws_const.protection.sheet = True
    ws_const.protection.enable()

    # handedness dropdown
    ws_site = sheets["Site Settings"]
    dv = DataValidation(type="list", formula1='"LHT,RHT"', allow_blank=False)
    ws_site.add_data_validation(dv)
    for row in ws_site.iter_rows(min_row=2):
        if row[0].value == "traffic_handedness":
            dv.add(row[1])

    # ---- Scenarios ----
    ws = wb.create_sheet("Scenarios")
    ws.append(["tag", "type", "vut_min_kmh", "vut_max_kmh", "side_impact", "has_sov"])
    _style_header(ws, 6)
    _autosize(ws, [12, 14, 12, 12, 12, 9])
    dv_type = DataValidation(type="list", formula1='"longitudinal,crossing,head-on"', allow_blank=False)
    ws.add_data_validation(dv_type)
    for tag, proto in raw.get("scenarios", {}).items():
        speed = proto.get("vut_speed_range_kmh") or [None, None]
        ws.append([
            tag, proto.get("type", "longitudinal"), speed[0], speed[1],
            bool(proto.get("side_impact", False)), bool(proto.get("has_sov", False)),
        ])
        dv_type.add(ws.cell(row=ws.max_row, column=2))

    # ---- Vehicle Dimensions ----
    ws = wb.create_sheet("Vehicle Dimensions")
    ws.append(["entity", "length_m", "width_m"])
    _style_header(ws, 3)
    _autosize(ws, [20, 12, 12])
    for name, dims in raw.get("vehicle_dimensions", {}).items():
        ws.append([name, dims.get("length"), dims.get("width")])

    # ---- Curve Radii ----
    ws = wb.create_sheet("Curve Radii")
    ws.append(["vut_speed_max_kmh", "direction", "radius_m"])
    _style_header(ws, 3)
    _autosize(ws, [20, 14, 12])
    for entry in raw.get("curve_part2_radii_m", []):
        ws.append([entry["vut_speed_max_kmh"], entry["direction"], entry["radius_m"]])

    # ---- Sim Time Bands ----
    ws = wb.create_sheet("Sim Time Bands")
    ws.append(["vut_speed_max_kmh", "min_s", "max_s"])
    _style_header(ws, 3)
    _autosize(ws, [20, 10, 10])
    for entry in raw.get("simulation_time_by_speed_s", []):
        ws.append([entry["vut_speed_max_kmh"], entry["min_s"], entry["max_s"]])

    return wb


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--json", default=str(_ROOT / "config.json"))
    parser.add_argument("--out", default=str(_ROOT / "config.xlsx"))
    args = parser.parse_args()

    raw = json.loads(Path(args.json).read_text())
    for key in [k for k in raw if k.startswith("_")]:
        raw.pop(key)

    wb = build_workbook(raw)
    wb.save(args.out)
    print(f"Wrote {args.out}")

    # Round-trip sanity check: both files must produce the identical Config.
    sys.path.insert(0, str(_ROOT))
    from src.models import Config
    a = Config.load(Path(args.json)).model_dump()
    b = Config.load(Path(args.out)).model_dump()
    # prefix ORDER is irrelevant (prefix matching) and the in-JSON 'description'
    # doc note has no behavioural effect - normalise both before comparing.
    # validation_column_widths is a JSON-only report-layout setting (not surfaced in the
    # Excel sheets), so it would always differ; drop it from the comparison.
    for d in (a, b):
        d["naming_convention"] = sorted(d["naming_convention"].get("valid_prefixes", []))
        d.pop("validation_column_widths", None)
    if a != b:
        for field in a:
            if a[field] != b[field]:
                print(f"  MISMATCH in '{field}'", file=sys.stderr)
        print("ERROR: config.xlsx does not round-trip to the same Config as config.json", file=sys.stderr)
        return 1
    print("Round-trip check OK: config.xlsx loads identically to config.json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
